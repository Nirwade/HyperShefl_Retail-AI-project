"""
export_ai.py — Export Center AI intelligence
Python prepares exact filtered exports, LLaMA explains what to download
"""
import requests
import pandas as pd
import io

def _llm(prompt):
    try:
        r = requests.post(
            "http://127.0.0.1:11434/api/generate",
            json={"model":"llama3.2:3b","prompt":prompt,"stream":False,
                  "options":{"temperature":0.1,"num_predict":180}},
            timeout=25)
        return r.json().get("response","").strip() if r.status_code==200 else None
    except:
        return None

def answer(question, na, sc, ns):
    """
    Returns (text, csv_bytes_or_None, filename, excel_bytes_or_None)
    """
    q = question.lower()
    NL = "\n"

    # ── Buyer report ──────────────────────────────────────────
    if any(x in q for x in ["buyer","report","buyer report","which file","what file","download"]):
        if na is None or na.empty:
            return "Alert data not available.", None, None, None

        crit = na[na["alert_tier"]=="CRITICAL"].copy() if "alert_tier" in na.columns else na.copy()
        crit = crit.sort_values("revenue_at_risk", ascending=False) if "revenue_at_risk" in crit.columns else crit

        rows = len(crit)
        stores = crit["store_id"].nunique() if "store_id" in crit.columns else 0
        rar = float(crit["revenue_at_risk"].sum()) if "revenue_at_risk" in crit.columns else 0

        facts = (
            f"Buyer report file: network_master_alerts_CRITICAL.csv\n"
            f"Filter: alert_tier = CRITICAL only\n"
            f"Rows: {rows:,} | Stores: {stores} | Revenue at risk: ${rar/1e6:.1f}M\n"
            f"Columns: store, SKU, product name, category, days left, units to order, "
            f"revenue at risk, supplier, priority score\n"
            f"Sorted by: revenue at risk descending (most urgent first)"
        )
        prompt = (
            "You are a retail data analyst.\n"
            "EXACT FILE DATA:\n" + facts + "\n\n"
            "Write 2-3 sentences as one paragraph. No bullet points. No headings.\n"
            "Tell the buyer exactly what this file contains, what filter is applied, "
            "and what to do with it. Be specific with the row count and revenue."
        )
        text = _llm(prompt) or facts
        csv_bytes = crit.to_csv(index=False).encode()
        return text, csv_bytes, "buyer_report_CRITICAL.csv", None

    # ── Region filtered CRITICAL ──────────────────────────────
    if any(x in q for x in ["region","southwest","west","northeast","southeast","my region","critical alert"]):
        if na is None or na.empty:
            return "Alert data not available.", None, None, None

        regions = na["region"].dropna().unique().tolist() if "region" in na.columns else []

        # Detect which region was mentioned
        region_found = None
        for reg in regions:
            if reg.lower() in q:
                region_found = reg
                break

        if region_found:
            filtered = na[(na["alert_tier"]=="CRITICAL") & (na["region"]==region_found)].copy() if "alert_tier" in na.columns else na[na["region"]==region_found].copy()
            rows = len(filtered)
            rar = float(filtered["revenue_at_risk"].sum()) if "revenue_at_risk" in filtered.columns else 0

            facts = (
                f"{region_found} CRITICAL alerts: {rows:,} rows, ${rar/1e6:.1f}M at risk\n"
                f"File: {region_found.lower().replace(' ','_')}_critical_alerts.csv"
            )
            prompt = (
                "You are a retail operations analyst.\n" + facts + "\n\n"
                "Write 2 sentences. Tell the buyer what this filtered file contains "
                f"and what immediate action to take for {region_found}. Use exact numbers."
            )
            text = _llm(prompt) or facts
            csv_bytes = filtered.to_csv(index=False).encode()
            fname = f"{region_found.lower().replace(' ','_')}_critical_alerts.csv"
            return text, csv_bytes, fname, None
        else:
            # List available regions
            region_list = ", ".join(regions)
            crit = na[na["alert_tier"]=="CRITICAL"] if "alert_tier" in na.columns else na
            reg_summary = []
            for reg in regions:
                sub = crit[crit["region"]==reg] if "region" in crit.columns else pd.DataFrame()
                rar = float(sub["revenue_at_risk"].sum()) if not sub.empty and "revenue_at_risk" in sub.columns else 0
                reg_summary.append(f"  {reg}: {len(sub):,} critical alerts · ${rar/1e6:.1f}M at risk")
            facts = "Available regions:\n" + NL.join(reg_summary)
            return f"Specify a region to filter. Available: {region_list}\n\n" + facts, None, None, None

    # ── Supplier scorecard Excel ──────────────────────────────
    if any(x in q for x in ["supplier","scorecard","excel","xlsx","supplier excel"]):
        if sc is None or sc.empty:
            return "Supplier scorecard not available.", None, None, None

        # Build Excel with risk tier color coding
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Supplier Scorecard"

            # Colors
            fills = {
                "HIGH_RISK":   PatternFill("solid", fgColor="FF4444"),
                "MEDIUM_RISK": PatternFill("solid", fgColor="F59E0B"),
                "LOW_RISK":    PatternFill("solid", fgColor="10B981"),
            }
            hdr_fill = PatternFill("solid", fgColor="0D9488")
            hdr_font = Font(bold=True, color="FFFFFF", size=10)
            white_font = Font(color="FFFFFF", size=9)

            # Display columns
            show_cols = [c for c in [
                "supplier_name","risk_tier","risk_score","avg_fulfillment_rate",
                "late_delivery_rate","short_delivery_rate","avg_lead_actual",
                "stockout_events_caused","total_revenue_at_risk","total_orders",
                "stores_served","skus_supplied"
            ] if c in sc.columns]

            col_labels = {
                "supplier_name":"Supplier","risk_tier":"Risk Tier","risk_score":"Score",
                "avg_fulfillment_rate":"Fill Rate","late_delivery_rate":"Late %",
                "short_delivery_rate":"Short %","avg_lead_actual":"Avg Lead Days",
                "stockout_events_caused":"Stockouts","total_revenue_at_risk":"Rev at Risk",
                "total_orders":"Orders","stores_served":"Stores","skus_supplied":"SKUs"
            }

            # Header row
            for ci,col in enumerate(show_cols, 1):
                cell = ws.cell(row=1, column=ci, value=col_labels.get(col, col))
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            sc_sorted = sc.sort_values("risk_score", ascending=False) if "risk_score" in sc.columns else sc
            for ri, (_, row) in enumerate(sc_sorted[show_cols].iterrows(), 2):
                tier = str(row.get("risk_tier","")) if "risk_tier" in show_cols else ""
                row_fill = fills.get(tier, None)
                for ci, col in enumerate(show_cols, 1):
                    val = row[col]
                    # Format percentages
                    if col in ["avg_fulfillment_rate","late_delivery_rate","short_delivery_rate"]:
                        val = f"{float(val)*100:.1f}%" if pd.notna(val) else ""
                    elif col == "total_revenue_at_risk":
                        val = f"${float(val):,.0f}" if pd.notna(val) else ""
                    elif col == "avg_lead_actual":
                        val = f"{float(val):.1f}" if pd.notna(val) else ""
                    cell = ws.cell(row=ri, column=ci, value=val)
                    if row_fill:
                        cell.fill = row_fill
                        cell.font = white_font

            # Column widths
            for ci, col in enumerate(show_cols, 1):
                ws.column_dimensions[get_column_letter(ci)].width = 16

            ws.freeze_panes = "A2"

            buf = io.BytesIO()
            wb.save(buf)
            excel_bytes = buf.getvalue()

            rows = len(sc)
            high = int((sc["risk_tier"]=="HIGH_RISK").sum()) if "risk_tier" in sc.columns else 0
            facts = (
                f"Supplier scorecard Excel: {rows} suppliers\n"
                f"HIGH RISK: {high} | Sorted by risk score descending\n"
                f"Color coded: red=HIGH, amber=MEDIUM, green=LOW"
            )
            prompt = (
                "You are a retail procurement analyst.\n" + facts + "\n\n"
                "Write 2 sentences. Tell the buyer what this Excel file contains "
                "and how to use the color coding to prioritize supplier reviews. Be specific."
            )
            text = _llm(prompt) or facts
            return text, None, None, excel_bytes

        except ImportError:
            return "openpyxl not installed. Run: pip install openpyxl --break-system-packages", None, None, None

    # ── Fallback ──────────────────────────────────────────────
    total_crit = int((na["alert_tier"]=="CRITICAL").sum()) if na is not None and not na.empty and "alert_tier" in na.columns else 0
    ctx = f"Network: {total_crit:,} CRITICAL alerts, {len(sc) if sc is not None else 0} suppliers."
    ans = _llm(
        "You are a retail data analyst helping download files.\n"
        "Context: " + ctx + "\n"
        "Answer in 2 sentences. Stay in retail data export scope.\n"
        "Question: " + question
    )
    return ans or "Try: buyer report · CRITICAL alerts by region · supplier scorecard Excel", None, None, None