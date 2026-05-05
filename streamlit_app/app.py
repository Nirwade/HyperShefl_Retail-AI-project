"""
HyperShelf v2.0 — DemandSense + Nexus — Streamlit Dashboard
All 4 bugs fixed:
1. fillcolor uses rgba() not hex+alpha
2. OLLAMA_HOST set for AI Assistant
3. Phantom uses is_phantom_candidate
4. ROI uses revenue_recovered column
"""
import os, sys, warnings
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
from pathlib import Path
from datetime import datetime
from scipy import stats as scipy_stats

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ["OLLAMA_HOST"] = "http://127.0.0.1:11434"

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PROCESSED    = PROJECT_ROOT / "data/processed/training"
NEXUS        = PROJECT_ROOT / "data/processed/nexus"
CSV          = PROJECT_ROOT / "data/raw/output/csv"

st.set_page_config(page_title="HyperShelf v2.0", layout="wide", initial_sidebar_state="expanded")



C_NAVY="#0F1E35";C_TEAL="#0D9488";C_TEAL2="#14B8A6"
C_RED="#EF4444";C_AMBER="#F59E0B";C_GREEN="#10B981"
C_PURPLE="#8B5CF6";C_WHITE="#FFFFFF";C_GRAY="#94A3B8"
C_CARD="#1E3352";C_BG="#0D1B2A";C_BORDER="#2D4A6A"
C_GOLD="#F59E0B";C_CYAN="#06B6D4"
TIER_COLORS={"CRITICAL":C_RED,"WARNING":C_AMBER,"MONITOR":C_PURPLE,"OK":C_GREEN}

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700;800&display=swap');
html,body,[class*="css"],*{{font-family:'IBM Plex Sans',sans-serif!important;}}
.stApp{{background:{C_BG}!important;}}
.stAppHeader{{display:none!important;}}
section[data-testid="stSidebar"]{{background:{C_NAVY}!important;border-right:1px solid {C_BORDER};min-width:220px!important;max-width:220px!important;}}
section[data-testid="stSidebar"],section[data-testid="stSidebar"] *,section[data-testid="stSidebar"] p,section[data-testid="stSidebar"] span,section[data-testid="stSidebar"] div,section[data-testid="stSidebar"] label{{color:{C_WHITE}!important;}}
section[data-testid="stSidebar"] .stRadio>div{{gap:0px!important;}}
section[data-testid="stSidebar"] .stRadio>div>label{{background:transparent!important;border:none!important;border-radius:6px!important;padding:9px 12px!important;margin:1px 0!important;display:flex!important;align-items:center!important;cursor:pointer!important;color:{C_GRAY}!important;font-size:13px!important;font-weight:500!important;width:100%!important;transition:all 0.15s!important;}}
section[data-testid="stSidebar"] .stRadio>div>label:hover{{background:{C_CARD}!important;color:{C_WHITE}!important;}}
section[data-testid="stSidebar"] .stRadio>div>label[data-checked="true"]{{background:{C_CARD}!important;border-left:3px solid {C_TEAL}!important;color:{C_TEAL2}!important;font-weight:700!important;}}
section[data-testid="stSidebar"] .stRadio>div>label>div:first-child{{display:none!important;}}
section[data-testid="stSidebar"] .stRadio>label{{display:none!important;}}
[data-testid="stSidebarCollapseButton"]{{display:none!important;}}
.block-container{{padding:0 1.5rem 4rem!important;max-width:100%!important;background:{C_BG};}}
.stTabs [data-baseweb="tab-list"]{{background:{C_NAVY};border:1px solid {C_BORDER};border-radius:6px;gap:2px;padding:3px;}}
.stTabs [data-baseweb="tab"]{{background:transparent;color:{C_GRAY};border-radius:4px;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;padding:7px 16px;}}
.stTabs [aria-selected="true"]{{background:{C_CARD}!important;color:{C_TEAL2}!important;}}
div[data-baseweb="select"]>div{{background:{C_NAVY}!important;border-color:{C_BORDER}!important;color:{C_WHITE}!important;}}
div[data-baseweb="menu"]{{background:{C_NAVY}!important;border-color:{C_BORDER}!important;}}
div[role="option"]{{color:{C_WHITE}!important;background:{C_NAVY}!important;}}
div[role="option"]:hover{{background:{C_CARD}!important;}}

/* ── DROPDOWN: selected value visible ── */
div[data-baseweb="select"] div[class*="placeholder"]{{color:{C_GRAY}!important;}}
div[data-baseweb="select"] div[class*="singleValue"]{{color:{C_WHITE}!important;}}
div[data-baseweb="select"] div[class*="ValueContainer"]{{color:{C_WHITE}!important;}}
div[data-baseweb="select"] div[class*="control"]{{background:{C_NAVY}!important;color:{C_WHITE}!important;}}
div[data-baseweb="select"] [class*="Input"]{{color:{C_WHITE}!important;}}
div[data-baseweb="select"] span{{color:{C_WHITE}!important;}}
div[data-baseweb="select"] p{{color:{C_WHITE}!important;}}

/* ── MULTISELECT tags ── */
span[data-baseweb="tag"]{{background:{C_TEAL}!important;color:{C_WHITE}!important;}}
span[data-baseweb="tag"] span{{color:{C_WHITE}!important;}}

/* ── SELECTBOX label + value ── */
div[data-testid="stSelectbox"] div[data-baseweb="select"] div{{color:{C_WHITE}!important;}}
div[data-testid="stMultiSelect"] div[data-baseweb="select"] div{{color:{C_WHITE}!important;}}

/* ── SLIDER value label ── */
div[data-testid="stSlider"] div[data-testid="stMarkdownContainer"] p{{color:{C_WHITE}!important;}}
[data-baseweb="select"] [data-testid="stMarkdownContainer"] p{{color:{C_WHITE}!important;}}
.stDataFrame table{{color:{C_WHITE};}}
.stDataFrame thead tr th{{background:{C_NAVY};color:{C_TEAL2};font-size:11px;text-transform:uppercase;border:1px solid {C_BORDER};}}
.stDataFrame tbody tr td{{background:{C_CARD};border:1px solid {C_BORDER};font-size:12px;color:{C_WHITE};}}
h1,h2,h3,h4,h5,h6{{color:{C_WHITE}!important;}}
p,span,div,label{{color:{C_WHITE};}}
hr{{border-color:{C_BORDER}!important;margin:16px 0!important;}}
.stButton>button{{background:{C_CARD};color:{C_TEAL2};border:1px solid {C_BORDER};border-radius:6px;font-size:11px;font-weight:700;}}
.stButton>button:hover{{background:{C_NAVY};border-color:{C_TEAL};}}
@keyframes ticker{{0%{{transform:translateX(0);}}100%{{transform:translateX(-33.33%);}}}}
.ticker-wrap{{background:{C_NAVY};border-bottom:1px solid {C_BORDER};padding:7px 0;overflow:hidden;margin:-1rem -1.5rem 1.5rem -1.5rem;}}
.ticker-move{{display:flex;align-items:center;white-space:nowrap;animation:ticker 50s linear infinite;}}
.ticker-item{{margin:0 24px;font-size:11px;}}
.kpi-row{{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px;}}
.kpi-card{{background:{C_CARD};border:1px solid {C_BORDER};border-radius:8px;padding:16px;flex:1;min-width:130px;}}
.kpi-title{{font-size:10px;font-weight:600;color:{C_GRAY};text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;}}
.kpi-val{{font-size:26px;font-weight:800;letter-spacing:-1px;line-height:1.1;}}
.kpi-sub{{font-size:10px;color:{C_GRAY};margin-top:4px;}}
.sh{{border-left:3px solid {C_TEAL};padding-left:12px;margin-bottom:16px;}}
.sh-title{{color:{C_WHITE};font-weight:700;font-size:14px;margin:0 0 3px 0;}}
.sh-sub{{color:{C_GRAY};font-size:11px;margin:0;}}
.card{{background:{C_CARD};border:1px solid {C_BORDER};border-radius:8px;padding:20px;margin-bottom:16px;}}
.card-glow-r{{background:{C_CARD};border:1px solid rgba(239,68,68,0.33);border-radius:8px;padding:20px;margin-bottom:16px;box-shadow:0 0 20px rgba(239,68,68,0.20);}}
.card-glow-g{{background:{C_CARD};border:1px solid rgba(16,185,129,0.33);border-radius:8px;padding:20px;margin-bottom:16px;box-shadow:0 0 20px rgba(16,185,129,0.20);}}
.card-glow-t{{background:{C_CARD};border:1px solid rgba(13,148,136,0.33);border-radius:8px;padding:20px;margin-bottom:16px;box-shadow:0 0 20px rgba(13,148,136,0.20);}}
.chat-ai{{background:{C_NAVY};border:1px solid {C_BORDER};border-radius:10px 10px 10px 2px;padding:10px 14px;font-size:12px;line-height:1.6;color:{C_WHITE};margin-bottom:8px;}}
.chat-user{{background:{C_TEAL};color:#fff;border-radius:10px 10px 2px 10px;padding:10px 14px;font-size:12px;line-height:1.6;margin-bottom:8px;text-align:right;}}
.chat-label{{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:{C_TEAL2};margin-bottom:3px;}}
::-webkit-scrollbar{{width:4px;height:4px;}}
::-webkit-scrollbar-track{{background:{C_BG};}}
::-webkit-scrollbar-thumb{{background:{C_BORDER};border-radius:3px;}}
.stDownloadButton>button{{background:{C_CARD}!important;color:{C_TEAL2}!important;border:1px solid {C_BORDER}!important;border-radius:6px!important;font-size:11px!important;font-weight:700!important;width:100%!important;}}
.stDownloadButton>button:hover{{background:{C_NAVY}!important;border-color:{C_TEAL}!important;}}
[data-testid="stMetric"]{{background:{C_CARD};border:1px solid {C_BORDER};border-radius:8px;padding:12px;}}
[data-testid="stMetricLabel"]{{color:{C_GRAY}!important;}}
[data-testid="stMetricValue"]{{color:{C_WHITE}!important;}}
[data-baseweb="tag"]{{background:{C_TEAL}!important;color:{C_WHITE}!important;}}
</style>
""", unsafe_allow_html=True)


def DL(title="", height=320, showlegend=True, **kw):
    base = dict(
        title=dict(text=title, font=dict(color=C_WHITE, size=13)),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color=C_WHITE, size=11), height=height,
        margin=dict(l=50, r=20, t=44 if title else 20, b=40),
        xaxis=dict(gridcolor=C_BORDER, zerolinecolor=C_BORDER, color=C_GRAY),
        yaxis=dict(gridcolor=C_BORDER, zerolinecolor=C_BORDER, color=C_GRAY),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=C_GRAY)),
        showlegend=showlegend,
        hoverlabel=dict(bgcolor=C_CARD, bordercolor=C_BORDER,
                        font=dict(family="IBM Plex Sans", color=C_WHITE, size=11)),
    )
    base.update(kw)
    return base

def EF(msg="No data available", height=320):
    fig = go.Figure()
    fig.add_annotation(text=msg, xref="paper", yref="paper", x=0.5, y=0.5,
                       showarrow=False, font=dict(color=C_GRAY, size=13))
    fig.update_layout(**DL(height=height))
    return fig

def PC(fig, key, height=None):
    if height: fig.update_layout(height=height)
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False}, key=key)

def kpi_row(cards):
    html_parts = ['<div class="kpi-row">']
    for title, value, subtitle, color in cards:
        html_parts.append(f'<div class="kpi-card" style="border-top:3px solid {color}"><div class="kpi-title">{title}</div><div class="kpi-val" style="color:{color}">{value}</div><div class="kpi-sub">{subtitle}</div></div>')
    html_parts.append("</div>")
    st.markdown("".join(html_parts), unsafe_allow_html=True)

def sh(title, subtitle=""):
    st.markdown(f'<div class="sh"><div class="sh-title">{title}</div>' + (f'<div class="sh-sub">{subtitle}</div>' if subtitle else "") + "</div>", unsafe_allow_html=True)


@st.cache_data(ttl=300, show_spinner=False)
def load_all():
    def rd(p, **kw):
        try: return pd.read_csv(p, **kw)
        except: return pd.DataFrame()
    d = {}
    d["stores"]    = rd(CSV/"stores.csv")
    d["products"]  = rd(CSV/"products.csv")
    d["suppliers"] = rd(CSV/"suppliers.csv")
    d["repl_logs"] = rd(CSV/"replenishment_logs.csv", parse_dates=["order_date","receive_date"])
    pred = rd(PROCESSED/"demandSense_v2_predictions.csv", parse_dates=["date"])
    if not pred.empty and "forecast_units" in pred.columns and "forecast_p50" not in pred.columns:
        pred["forecast_p50"] = pred["forecast_units"]
    d["pred"]       = pred
    d["repl"]       = rd(PROCESSED/"replenishment_policy_inputs_demandsense.csv")
    d["alerts_ml"]  = rd(PROCESSED/"weekly_alerts.csv")
    d["weekly"]     = rd(PROCESSED/"weekly_monitor_demandsense.csv")
    d["model_sum"]  = rd(PROCESSED/"demandSense_model_summary.csv")
    d["cat_impact"] = rd(PROCESSED/"category_impact_scores.csv")
    d["backtest_s"] = rd(PROCESSED/"backtest_summary_demandsense.csv")
    ns = rd(NEXUS/"allstore"/"network_store_summary.csv")
    if not ns.empty:
        ns = ns.rename(columns={"critical":"critical_count","warning":"warning_count",
                                 "monitor":"monitor_count","ok":"ok_count",
                                 "phantom_count":"phantom_skus","total_skus":"skus_analyzed"})
        stores = d["stores"]
        for col in ["store_name","city","state","region","store_format","foot_traffic_tier"]:
            if col not in ns.columns and col in stores.columns:
                ns = ns.merge(stores[["store_id",col]], on="store_id", how="left")
    d["ns"]         = ns
    d["na"]         = rd(NEXUS/"allstore"/"network_master_alerts.csv")
    d["phantom_c"]  = rd(NEXUS/"allstore"/"phantom_confidence.csv")
    d["phantom_s"]  = rd(NEXUS/"allstore"/"phantom_by_store.csv")
    d["loc_scores"] = rd(NEXUS/"localization"/"store_localization_scores.csv")
    d["loc_cat"]    = rd(NEXUS/"localization"/"store_category_mismatch.csv")
    d["loc_net"]    = rd(NEXUS/"localization"/"category_network_mismatch.csv")
    d["sup_sc"]     = rd(NEXUS/"supplier"/"supplier_scorecard.csv")
    d["roi_sum"]    = rd(NEXUS/"roi"/"roi_executive_summary.csv")
    d["roi_reg"]    = rd(NEXUS/"roi"/"roi_by_region.csv")
    d["rc"]         = rd(NEXUS/"forecast"/"stockout_by_cause.csv")
    d["acc_cat"]    = rd(NEXUS/"forecast"/"accuracy_by_category.csv")
    d["acc_month"]  = rd(NEXUS/"forecast"/"accuracy_by_month.csv")
    d["report_time"]= datetime.now().strftime("%A, %B %d, %Y at %I:%M %p")
    return d


@st.cache_resource
def get_agent(persona):
    try:
        from src.agent import DemandSenseAgent
        return DemandSenseAgent(persona=persona)
    except: return None


D  = load_all()
ns = D["ns"]; na = D["na"]
STORES = D["stores"]["store_id"].tolist() if not D["stores"].empty else []
S_NAMES = {}
if not D["stores"].empty:
    for _, r in D["stores"].iterrows():
        S_NAMES[r["store_id"]] = f"{r['store_id']} — {r.get('store_name','')}"

REGIONS = ["All"] + (sorted(ns["region"].dropna().unique().tolist()) if not ns.empty and "region" in ns.columns else [])
FORMATS = ["All"] + (sorted(ns["store_format"].dropna().unique().tolist()) if not ns.empty and "store_format" in ns.columns else [])
TIERS_F = ["All"] + (sorted(ns["foot_traffic_tier"].dropna().unique().tolist()) if not ns.empty and "foot_traffic_tier" in ns.columns else [])

for k in ["chat","page"]:
    if k not in st.session_state:
        st.session_state[k] = {} if k=="chat" else "Executive Overview"


def render_ticker():
    def g(col, d=0): return float(ns[col].sum()) if not ns.empty and col in ns.columns else d
    rar=g("revenue_at_risk"); crit=int(g("critical_count")); warn=int(g("warning_count")); ph=int(g("phantom_skus"))
    items = [
        f'<span style="color:{C_RED};font-weight:700">CRITICAL</span> <span style="color:{C_WHITE}">{crit:,} SKUs need immediate order</span>',
        f'<span style="color:{C_AMBER};font-weight:700">REVENUE AT RISK</span> <span style="color:{C_WHITE}">${rar/1e6:.1f}M network-wide today</span>',
        f'<span style="color:{C_AMBER};font-weight:700">WARNING</span> <span style="color:{C_WHITE}">{warn:,} SKUs order today</span>',
        f'<span style="color:{C_PURPLE};font-weight:700">PHANTOM</span> <span style="color:{C_WHITE}">{ph} ghost SKUs detected</span>',
        f'<span style="color:{C_GRAY};font-weight:700">LAST UPDATE</span> <span style="color:{C_WHITE}">{D["report_time"]}</span>',
        f'<span style="color:{C_TEAL2};font-weight:700">STORES</span> <span style="color:{C_WHITE}">{len(STORES)} monitored continuously</span>',
    ]
    inner = "".join(f'<span class="ticker-item">{it}</span>' for it in items)
    st.markdown(f'<div class="ticker-wrap"><div class="ticker-move">{inner*3}</div></div>', unsafe_allow_html=True)


def chat_panel(page_key, persona, intro, quick=None, page_context=None):
    st.markdown("<hr>", unsafe_allow_html=True)
    sh("AI Assistant", f"Persona: {persona.replace('_',' ').title()}")
    if page_key not in st.session_state["chat"]:
        st.session_state["chat"][page_key] = []
    msgs = st.session_state["chat"][page_key]
    if quick:
        cols = st.columns(len(quick))
        for i,(lbl,q) in enumerate(quick):
            with cols[i]:
                if st.button(lbl, key=f"qq_{page_key}_{i}", use_container_width=True):
                    st.session_state[f"_pend_{page_key}"] = q
    html = ""
    if not msgs:
        html = f'<div class="chat-label">DemandSense AI</div><div class="chat-ai">{intro}</div>'
    for m in msgs:
        if m["role"] == "user":
            html += f'<div class="chat-label" style="text-align:right">You</div><div class="chat-user">{m["content"]}</div>'
        else:
            txt = m["content"].replace("<","&lt;").replace(">","&gt;").replace("\n","<br>")
            html += f'<div class="chat-label">AI</div><div class="chat-ai">{txt}</div>'
    if html: st.markdown(html, unsafe_allow_html=True)
    pending = st.session_state.pop(f"_pend_{page_key}", None)
    user_in = st.chat_input("Ask anything...", key=f"ci_{page_key}")
    user_in = user_in or pending
    if user_in:
        msgs.append({"role":"user","content":user_in})
        agent = get_agent(persona)
        with st.spinner("Thinking..."):
            if agent:
                try:
                    ctx = []
                    if not ns.empty:
                        rar2  = float(ns["revenue_at_risk"].sum()) if "revenue_at_risk" in ns.columns else 0
                        crit2 = int(ns["critical_count"].sum()) if "critical_count" in ns.columns else 0
                        warn2 = int(ns["warning_count"].sum()) if "warning_count" in ns.columns else 0
                        ctx.append(f"LIVE DATA: Revenue at risk ${rar2:,.0f}, Critical SKUs {crit2:,}, Warning SKUs {warn2:,}, Stores {len(ns)}")
                        if "store_id" in ns.columns and "revenue_at_risk" in ns.columns:
                            top3 = ns.nlargest(3,"revenue_at_risk")[["store_id","critical_count","revenue_at_risk"]].to_dict("records")
                            ctx.append(f"TOP 3 URGENT STORES: {top3}")
                    if page_context:
                        ctx.append(f'PAGE DATA:\n{page_context}')
                    ctx.append('IMPORTANT: Answer ONLY using the data above. Do not invent any numbers, order IDs, or store names not in the data provided.')
                    grounded = ("\n\nCONTEXT:\n" + "\n".join(ctx) + "\n\nQUESTION: " + user_in) if ctx else user_in
                    r = agent.chat(grounded)
                    resp = r.get("response","") if isinstance(r,dict) else str(r)
                    if not resp or len(resp.strip()) < 5:
                        resp = "I processed your request. Try asking about a specific store or network KPI."
                    if isinstance(r,dict) and r.get("requires_approval"):
                        resp += f"\n\nApproval required: {r.get('escalation_reason','')}"
                except Exception as e:
                    resp = f"Error: {e}"
            else:
                resp = "AI not available — run: ollama serve"
        msgs.append({"role":"assistant","content":resp})
        st.rerun()


# ── SIDEBAR ──────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f'<div style="padding:12px 8px 20px;display:flex;align-items:center;gap:10px"><div style="font-size:26px;color:{C_TEAL}">&#x2B21;</div><div><div style="font-size:15px;font-weight:800;color:{C_WHITE}">HyperShelf</div><div style="font-size:8px;color:{C_TEAL2};text-transform:uppercase;letter-spacing:1.5px">v2.0 · Retail Intelligence</div></div></div>', unsafe_allow_html=True)
    PAGES = ["Executive Overview","Store Deep Dive","What-If Simulator","Store Comparison",
             "Safety Stock & ROP","Phantom Inventory","Localization","Supplier Performance",
             "Forecast & Model","Export Center"]
    selected = st.radio("Navigation", PAGES, label_visibility="collapsed", key="nav_radio")
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown(f'<div style="font-size:9px;color:{C_GRAY};text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:6px">Global Store</div>', unsafe_allow_html=True)
    g_store_sel = st.selectbox("Global Store", ["All stores"]+[S_NAMES.get(s,s) for s in STORES[:80]], label_visibility="collapsed", key="g_store")
    g_store_id  = None if g_store_sel=="All stores" else g_store_sel.split(" — ")[0]
    if g_store_id and not D["stores"].empty:
        sr = D["stores"][D["stores"]["store_id"]==g_store_id]
        if not sr.empty:
            r = sr.iloc[0]
            st.markdown(f'<div style="font-size:11px;color:{C_GRAY};padding:6px 4px;line-height:1.8">{r.get("city","")}, {r.get("state","")}<br>{r.get("store_format","")}<br>{r.get("foot_traffic_tier","")}</div>', unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)
    if st.button("Clear all chats", use_container_width=True, key="clear_chats"):
        st.session_state["chat"] = {}
        st.rerun()
    st.markdown('<div style="height:52px"></div>', unsafe_allow_html=True)
    st.markdown(f'<div style="position:absolute;bottom:16px;left:0;right:0;text-align:center"><div style="font-size:9px;color:{C_GRAY}">Nexus · Team 2</div><div style="font-size:9px;color:{C_GRAY}">UB + Globant AI Studios</div></div>', unsafe_allow_html=True)


PAGE = selected
render_ticker()


# ══════════════════════════════════════════════════════════════
# EXECUTIVE OVERVIEW
# ══════════════════════════════════════════════════════════════
if PAGE == "Executive Overview":
    sh("Executive Overview","Network KPIs · hotspot alert · region · format · ROI")
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Dynamic hotspot alert ─────────────────────────────────
    if not ns.empty and "region" in ns.columns:
        _worst_reg = ns.groupby("region")["revenue_at_risk"].sum().idxmax()
        _worst_rev = ns.groupby("region")["revenue_at_risk"].sum().max()
        _worst_crit = int(ns[ns["region"]==_worst_reg]["critical_count"].sum())
        _worst_stores = int(ns[ns["region"]==_worst_reg]["store_id"].nunique())
        st.markdown(f'<div style="background:rgba(239,68,68,0.10);border:1px solid rgba(239,68,68,0.40);border-radius:8px;padding:14px 20px;margin-bottom:20px;display:flex;align-items:center;gap:16px"><span style="font-size:20px">⚠️</span><div><div style="font-size:13px;font-weight:700;color:{C_RED}">HOTSPOT TODAY — {_worst_reg} region is most critical</div><div style="font-size:11px;color:{C_GRAY};margin-top:3px">${_worst_rev/1e6:.1f}M revenue at risk · {_worst_crit:,} critical SKUs · {_worst_stores} stores · Updated {D["report_time"]}</div></div></div>', unsafe_allow_html=True)

    # ── Filters ───────────────────────────────────────────────
    fc1,fc2,fc3,fc4 = st.columns([1.2,1.2,1.2,1.5])
    with fc1:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Region</div>', unsafe_allow_html=True)
        f_region = st.selectbox("r",REGIONS,label_visibility="collapsed",key="ov_region")
    with fc2:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Format</div>', unsafe_allow_html=True)
        f_format = st.selectbox("f",FORMATS,label_visibility="collapsed",key="ov_format")
    with fc3:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Traffic Tier</div>', unsafe_allow_html=True)
        f_tier = st.selectbox("t",TIERS_F,label_visibility="collapsed",key="ov_tier")
    with fc4:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Top N Stores</div>', unsafe_allow_html=True)
        f_topn = st.slider("n",5,30,10,5,label_visibility="collapsed",key="ov_topn")
    st.markdown("<hr>", unsafe_allow_html=True)

    ns_f = ns.copy()
    if not ns_f.empty:
        if f_region!="All" and "region" in ns_f.columns: ns_f=ns_f[ns_f["region"]==f_region]
        if f_format!="All" and "store_format" in ns_f.columns: ns_f=ns_f[ns_f["store_format"]==f_format]
        if f_tier!="All" and "foot_traffic_tier" in ns_f.columns: ns_f=ns_f[ns_f["foot_traffic_tier"]==f_tier]

    def g(col,d=0): return float(ns_f[col].sum()) if not ns_f.empty and col in ns_f.columns else d
    _roi=D["roi_sum"]
    roi_rec = float(_roi["revenue_recovered"].iloc[0]) if not _roi.empty and "revenue_recovered" in _roi.columns else 0
    pct_rec = float(_roi["pct_recovered"].iloc[0]) if not _roi.empty and "pct_recovered" in _roi.columns else 0

    # ── KPI strip with sparkline indicators ──────────────────
    kpi_row([
        ("Revenue at Risk", f"${g('revenue_at_risk')/1e6:.1f}M", "Network-wide today", C_RED),
        ("Critical SKUs", f"{int(g('critical_count')):,}", "Order immediately", C_RED),
        ("Warning SKUs", f"{int(g('warning_count')):,}", "Order today", C_AMBER),
        ("Units to Order", f"{int(g('units_to_order')):,}", "CRITICAL + WARNING", C_TEAL),
        ("Phantom SKUs", f"{int(g('phantom_skus')):,}", "Ghost inventory", C_PURPLE),
        ("Revenue Recovered", f"${roi_rec/1e6:.1f}M", f"{pct_rec:.1f}% of total lost", C_GREEN),
    ])

    # ── Region breakdown sparklines ───────────────────────────
    if not ns_f.empty and "region" in ns_f.columns:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        sh("Region Performance Breakdown","Revenue at risk · critical count · store count")
        rg_s = ns_f.groupby("region").agg(
            stores=("store_id","count"),
            critical=("critical_count","sum"),
            revenue=("revenue_at_risk","sum"),
            phantoms=("phantom_skus","sum")
        ).reset_index().sort_values("revenue",ascending=False)
        rg_s["rev_pct"] = rg_s["revenue"]/rg_s["revenue"].sum()*100

        rg_cols = st.columns(len(rg_s))
        for ci,(_, r) in enumerate(rg_s.iterrows()):
            with rg_cols[ci]:
                color = C_RED if r["rev_pct"]>30 else (C_AMBER if r["rev_pct"]>20 else C_GREEN)
                st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {color};border-radius:8px;padding:12px;text-align:center"><div style="font-size:11px;font-weight:700;color:{C_WHITE};margin-bottom:6px">{r["region"]}</div><div style="font-size:18px;font-weight:800;color:{color}">${r["revenue"]/1e6:.1f}M</div><div style="font-size:10px;color:{C_GRAY}">{int(r["critical"]):,} critical</div><div style="font-size:10px;color:{C_GRAY}">{int(r["stores"])} stores</div><div style="font-size:9px;color:{color};font-weight:700;margin-top:4px">{r["rev_pct"]:.1f}% of risk</div></div>', unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # ── Charts row ────────────────────────────────────────────
    r1c1,r1c2,r1c3 = st.columns(3)
    with r1c1:
        sh("Revenue at Risk by Region")
        if not ns_f.empty and "region" in ns_f.columns and "revenue_at_risk" in ns_f.columns:
            rr=ns_f.groupby("region")["revenue_at_risk"].sum().sort_values()
            fig=go.Figure(go.Bar(x=rr.values,y=rr.index,orientation="h",
                marker=dict(color=list(rr.values),colorscale=[[0,C_AMBER],[1,C_RED]],showscale=False),
                text=[f"${v/1e6:.1f}M" for v in rr.values],textposition="outside",
                textfont=dict(color=C_WHITE,size=9)))
            fig.update_layout(**DL("",height=280))
            PC(fig,"ov_rar_reg")
        else: PC(EF(),"ov_rar_reg_e")
    with r1c2:
        sh("Alert Tier Distribution")
        if not ns_f.empty:
            cols_m=[("CRITICAL","critical_count",C_RED),("WARNING","warning_count",C_AMBER),
                    ("MONITOR","monitor_count",C_PURPLE),("OK","ok_count",C_GREEN)]
            labels=[l for l,c,_ in cols_m if c in ns_f.columns]
            values=[int(ns_f[c].sum()) for _,c,_ in cols_m if c in ns_f.columns]
            colors=[cl for _,c,cl in cols_m if c in ns_f.columns]
            if values:
                fig2=go.Figure(go.Pie(labels=labels,values=values,hole=0.58,
                    marker_colors=colors,textinfo="label+percent",
                    textfont=dict(color=C_WHITE,size=10)))
                fig2.add_annotation(text=f"{sum(values):,}<br>SKUs",x=0.5,y=0.5,
                    showarrow=False,font=dict(color=C_WHITE,size=12))
                fig2.update_layout(**DL("",showlegend=False,height=280))
                PC(fig2,"ov_tier_pie")
            else: PC(EF(),"ov_tier_pie_e")
    with r1c3:
        sh("Critical Alerts by Store Format")
        if not ns_f.empty and "store_format" in ns_f.columns and "critical_count" in ns_f.columns:
            fc_g=ns_f.groupby("store_format")["critical_count"].sum().sort_values()
            fig3=go.Figure(go.Bar(x=fc_g.values,y=fc_g.index,orientation="h",
                marker_color=[C_RED if v==fc_g.max() else C_AMBER for v in fc_g.values],
                text=fc_g.values,textposition="outside",textfont=dict(color=C_WHITE,size=9)))
            fig3.update_layout(**DL("",height=280))
            PC(fig3,"ov_fmt")
        else: PC(EF(),"ov_fmt_e")

    r2c1,r2c2 = st.columns([1,2])
    with r2c1:
        sh("Safety Stock Calibration")
        if not ns_f.empty and all(c in ns_f.columns for c in ["understocked","overstocked","optimal"]):
            us=int(ns_f["understocked"].sum());ov=int(ns_f["overstocked"].sum());op=int(ns_f["optimal"].sum())
            fig_ss=go.Figure(go.Pie(labels=["Understocked","Overstocked","Optimal"],
                values=[us,ov,op],hole=0.55,marker_colors=[C_RED,C_AMBER,C_GREEN],
                textinfo="label+percent",textfont=dict(color=C_WHITE,size=10)))
            fig_ss.add_annotation(text=f"{us+ov+op:,}<br>SKUs",x=0.5,y=0.5,
                showarrow=False,font=dict(color=C_WHITE,size=12))
            fig_ss.update_layout(**DL("",showlegend=False,height=280))
            PC(fig_ss,"ov_ss")
        else: PC(EF("Run analytics.py"),"ov_ss_e")
    with r2c2:
        sh("Revenue at Risk vs Phantom SKUs","Bubble size = units to order")
        if not ns_f.empty and all(c in ns_f.columns for c in ["phantom_skus","revenue_at_risk","units_to_order"]):
            ns2=ns_f[ns_f["revenue_at_risk"]>0].copy()
            if not ns2.empty:
                mx=max(ns2["units_to_order"].max(),1)
                fig_b=go.Figure(go.Scatter(x=ns2["revenue_at_risk"],y=ns2["phantom_skus"],
                    mode="markers",
                    marker=dict(size=(ns2["units_to_order"]/mx*40).clip(4,40),
                        color=ns2["critical_count"] if "critical_count" in ns2.columns else ns2["revenue_at_risk"],
                        colorscale="Reds",showscale=True,
                        colorbar=dict(title=dict(text="Critical",font=dict(color=C_WHITE)),
                            tickfont=dict(color=C_WHITE))),
                    text=[f"{r['store_id']}<br>${r['revenue_at_risk']:,.0f}<br>{int(r['phantom_skus'])} phantoms"
                          for _,r in ns2.iterrows()],
                    hovertemplate="%{text}<extra></extra>"))
                fig_b.update_layout(**DL("",height=280))
                fig_b.update_xaxes(title="Revenue at Risk ($)")
                fig_b.update_yaxes(title="Phantom SKUs")
                PC(fig_b,"ov_bubble")
            else: PC(EF(),"ov_bubble_e")
        else: PC(EF("Run analytics.py first"),"ov_bubble_na")

    # ── ROI Simulator ─────────────────────────────────────────
    st.markdown('<div class="card-glow-g">', unsafe_allow_html=True)
    sh("ROI Simulator","How much revenue would HyperShelf recover? Drag the sliders.")
    sc1_,sc2_ = st.columns([2,1])
    with sc1_:
        roi_pct=st.slider("% of Critical orders actioned today",0,100,80,5,key="roi_pct")
        roi_sl=st.slider("Service level target",0.90,0.999,0.975,0.005,format="%.3f",key="roi_sl")
    with sc2_:
        total_rar_=float(ns_f["revenue_at_risk"].sum()) if not ns_f.empty and "revenue_at_risk" in ns_f.columns else 0
        new_z=scipy_stats.norm.ppf(roi_sl)
        z_ratio=new_z/1.96
        adj_rar=total_rar_*(roi_pct/100)
        st.markdown(f'<div style="text-align:center;padding:12px;background:{C_NAVY};border-radius:8px"><div style="font-size:30px;font-weight:800;color:{C_GREEN}">${adj_rar:,.0f}</div><div style="font-size:11px;color:{C_GRAY};margin-top:4px">Recoverable if {roi_pct}% of critical orders placed now</div><hr style="border-color:{C_BORDER};margin:10px 0"><div style="font-size:11px;color:{C_TEAL2};font-weight:600">Z = {new_z:.4f} ({roi_sl*100:.1f}% SL)</div><div style="font-size:10px;color:{C_AMBER};margin-top:4px">Safety stock impact: ×{z_ratio:.2f} vs current</div></div>', unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    # ── Top N stores table ────────────────────────────────────
    st.markdown('<div class="card">', unsafe_allow_html=True)
    sh(f"Top {f_topn} Urgent Stores","Sorted by urgency score")
    if not ns_f.empty:
        sort_c="urgency_score" if "urgency_score" in ns_f.columns else "revenue_at_risk"
        top_n=ns_f.nlargest(f_topn,sort_c)
        show_c=[c for c in ["store_id","store_name","city","state","region","store_format",
                             "critical_count","warning_count","revenue_at_risk",
                             "phantom_skus","urgency_score"] if c in top_n.columns]
        st.dataframe(top_n[show_c],use_container_width=True,height=300,hide_index=True)
    else:
        st.markdown(f'<div style="color:{C_GRAY};padding:20px;text-align:center">Run analytics.py to populate.</div>', unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    # ── Executive AI ──────────────────────────────────────────
    st.markdown("<hr>", unsafe_allow_html=True)
    sh("AI Assistant — Executive Intelligence","")
    try:
        import importlib, sys as _sys4
        _sp4=str(Path(__file__).resolve().parent)
        if _sp4 not in _sys4.path: _sys4.path.insert(0,_sp4)
        import executive_ai as _eai
        importlib.reload(_eai)
        _eai_ok=True
    except:
        _eai_ok=False
    _ev_key="overview"
    if _ev_key not in st.session_state["chat"]:
        st.session_state["chat"][_ev_key]=[]
    _ev_msgs=st.session_state["chat"][_ev_key]
    _ev_quick=[
        ("Worst region","Which region is worst right now and why?"),
        ("Format risk","Which store format has the most revenue at risk?"),
        ("Recover now","How much revenue do I recover if I order now?"),
        ("Premium stores","Show me Premium stores at risk"),
        ("Network KPIs","Give me a network KPI summary for today"),
    ]
    _evcols=st.columns(len(_ev_quick))
    for _evi,(_evl,_evq) in enumerate(_ev_quick):
        with _evcols[_evi]:
            if st.button(_evl,key=f"qq_ev_{_evi}",use_container_width=True):
                st.session_state["_pend_ev"]=_evq
    for _m in _ev_msgs:
        if _m["role"]=="user":
            _uh='<div class="chat-label" style="text-align:right">You</div>'
            _uh+='<div class="chat-user">'+str(_m["content"])+'</div>'
            st.markdown(_uh,unsafe_allow_html=True)
        else:
            _etxt=str(_m["content"]).replace("<","&lt;").replace(">","&gt;").replace(chr(10),"<br>")
            _ah='<div class="chat-label">AI</div><div class="chat-ai">'+_etxt+'</div>'
            st.markdown(_ah,unsafe_allow_html=True)
            if _m.get("fig") is not None:
                try:
                    st.plotly_chart(_m["fig"],use_container_width=True,
                        config={"displayModeBar":False},key="ev_chart_"+str(_ev_msgs.index(_m)))
                except: pass
    if not _ev_msgs:
        st.markdown('<div class="chat-ai">Ask about the worst region, store format risk, Premium stores, or how much revenue you recover if you order now.</div>',unsafe_allow_html=True)
    _ev_pend=st.session_state.pop("_pend_ev",None)
    _ev_in=st.chat_input("Ask about network performance...",key="ci_overview")
    _ev_in=_ev_in or _ev_pend
    if _ev_in:
        _ev_msgs.append({"role":"user","content":_ev_in,"fig":None})
        with st.spinner("Analysing network data..."):
            if _eai_ok:
                _etxt2,_efig2=_eai.answer(_ev_in,ns_f,D["roi_sum"])
            else:
                _etxt2="executive_ai module not loaded."
                _efig2=None
        _ev_msgs.append({"role":"assistant","content":_etxt2,"fig":_efig2})
        st.rerun()

elif PAGE == "Store Deep Dive":
    sh("Store Deep Dive","Live alerts · cost of inaction · order list · supplier intelligence")
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Store + filters ───────────────────────────────────────
    c1,c2,c3,c4 = st.columns([2,1.5,1.5,1.5])
    with c1:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Select Store</div>', unsafe_allow_html=True)
        store_sel=st.selectbox("ss",[S_NAMES.get(s,s) for s in STORES[:80]],label_visibility="collapsed",key="store_sel")
        store_id=store_sel.split(" — ")[0] if store_sel else (STORES[0] if STORES else None)
    with c2:
        avail_cats=sorted(na[na["store_id"]==store_id]["category"].dropna().unique().tolist()) if not na.empty and store_id and "category" in na.columns else []
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Category</div>', unsafe_allow_html=True)
        f_cats=st.multiselect("cats",avail_cats,label_visibility="collapsed",key="store_cats",placeholder="All categories")
    with c3:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Alert Tier</div>', unsafe_allow_html=True)
        f_atiers=st.multiselect("tiers",["CRITICAL","WARNING","MONITOR","OK"],label_visibility="collapsed",key="store_atiers",placeholder="All tiers")
    with c4:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Sort by</div>', unsafe_allow_html=True)
        sort_by=st.selectbox("sortby",["Priority Score","Days Left","Revenue at Risk"],label_visibility="collapsed",key="store_sortby")
    st.markdown("<hr>", unsafe_allow_html=True)

    if not na.empty and store_id:
        s=na[na["store_id"]==store_id].copy()
        if f_cats: s=s[s["category"].isin(f_cats)]
        if f_atiers: s=s[s["alert_tier"].isin(f_atiers)]
        sort_col_map={"Priority Score":"priority_score","Days Left":"days_of_supply_current","Revenue at Risk":"revenue_at_risk"}
        sort_col=sort_col_map.get(sort_by,"priority_score")
        if sort_col in s.columns: s=s.sort_values(sort_col,ascending=sort_col=="days_of_supply_current")

        si=ns[ns["store_id"]==store_id].iloc[0] if not ns.empty and store_id in ns["store_id"].values else pd.Series()
        def sv(col,d="—"): return si.get(col,d) if hasattr(si,"get") and col in si.index else d
        tc=s["alert_tier"].value_counts().to_dict() if "alert_tier" in s.columns else {}
        rar=float(s["revenue_at_risk"].sum()) if "revenue_at_risk" in s.columns else 0
        uto=float(s["units_to_order"].sum()) if "units_to_order" in s.columns else 0
        ph_n=int(s["is_phantom"].sum()) if "is_phantom" in s.columns else 0

        st.markdown(f'<div style="margin-bottom:16px"><div style="font-size:16px;font-weight:800;color:{C_WHITE}">{sv("store_name")} · {sv("city")}, {sv("state")}</div><div style="font-size:11px;color:{C_GRAY};margin-top:2px">{sv("region")} · {sv("store_format")} · {sv("foot_traffic_tier")}</div></div>', unsafe_allow_html=True)

        kpi_row([
            ("CRITICAL",str(tc.get("CRITICAL",0)),"Order NOW",C_RED),
            ("WARNING",str(tc.get("WARNING",0)),"Order today",C_AMBER),
            ("MONITOR",str(tc.get("MONITOR",0)),"Watch",C_PURPLE),
            ("OK",str(tc.get("OK",0)),"Well stocked",C_GREEN),
            ("Revenue at Risk",f"${rar:,.0f}","",C_RED),
            ("Units to Order",f"{uto:,.0f}","",C_TEAL),
            ("Phantom SKUs",str(ph_n),"Check aisles",C_PURPLE)])

        cc1,cc2=st.columns([1,2])
        with cc1:
            sh("Alert Tier Breakdown")
            fig_t=go.Figure(go.Pie(
                labels=["CRITICAL","WARNING","MONITOR","OK"],
                values=[int(tc.get(t,0)) for t in ["CRITICAL","WARNING","MONITOR","OK"]],
                hole=0.6,marker_colors=[C_RED,C_AMBER,C_PURPLE,C_GREEN],
                textinfo="label+value",textfont=dict(color=C_WHITE,size=10)))
            fig_t.update_layout(**DL("",height=260,showlegend=False))
            PC(fig_t,"store_tier")
        with cc2:
            sh("Revenue at Risk by Category")
            if "category" in s.columns and "revenue_at_risk" in s.columns:
                cr=s.groupby("category")["revenue_at_risk"].sum().sort_values(ascending=True)
                fig_c=go.Figure(go.Bar(x=cr.values,y=cr.index,orientation="h",
                    marker=dict(color=list(cr.values),colorscale=[[0,C_TEAL],[0.5,C_AMBER],[1,C_RED]],showscale=False),
                    text=[f"${v:,.0f}" for v in cr.values],textposition="outside",
                    textfont=dict(color=C_WHITE,size=8)))
                fig_c.update_layout(**DL("",height=260))
                PC(fig_c,"store_cat")
            else: PC(EF(height=260),"store_cat_e")

        # Daily cost of inaction
        st.markdown('<div class="card-glow-r">', unsafe_allow_html=True)
        sh("Daily Cost of Inaction","Revenue lost per day for every CRITICAL product with no order placed")
        crit_s=s[s["alert_tier"]=="CRITICAL"].copy() if "alert_tier" in s.columns else pd.DataFrame()
        if not crit_s.empty and "demand_for_calc" in crit_s.columns and "unit_price" in crit_s.columns:
            crit_s["daily_cost"]=crit_s["demand_for_calc"]*crit_s["unit_price"].fillna(0)
            total_daily=crit_s["daily_cost"].sum()
            st.markdown(f'<div style="font-size:28px;font-weight:800;color:{C_RED};margin-bottom:8px">${total_daily:,.0f} / day</div>', unsafe_allow_html=True)
            top5=crit_s.nlargest(5,"daily_cost")[["sku_id","product_name","daily_cost"]] if "product_name" in crit_s.columns else crit_s.nlargest(5,"daily_cost")[["sku_id","daily_cost"]]
            for _,r in top5.iterrows():
                st.markdown(f'<div style="display:flex;justify-content:space-between;margin-bottom:4px;padding:4px 0;border-bottom:1px solid {C_BORDER}30"><span style="color:{C_GRAY};font-size:10px;width:80px">{r.get("sku_id","")}</span><span style="color:{C_WHITE};font-size:11px;flex:1">{str(r.get("product_name",""))[:35]}</span><span style="color:{C_RED};font-weight:700;font-size:11px">${r["daily_cost"]:,.0f}/day</span></div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div style="font-size:24px;font-weight:800;color:{C_RED}">${rar/7:,.0f} / day</div>', unsafe_allow_html=True)
            st.markdown(f'<div style="color:{C_GRAY};font-size:12px">Estimated (revenue at risk / 7 days)</div>', unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        # Gauges
        if "alert_tier" in s.columns and "days_of_supply_current" in s.columns:
            top6=s[s["alert_tier"]=="CRITICAL"].head(6)
            if not top6.empty:
                st.markdown('<div class="card">', unsafe_allow_html=True)
                sh("Product Confidence Gauges — Top 6 Critical","How close is each product to stockout?")
                g_cols=st.columns(min(6,len(top6)))
                for i,(_,r) in enumerate(top6.iterrows()):
                    with g_cols[i]:
                        dos=float(r["days_of_supply_current"])
                        lt=float(r.get("lead_time_final",r.get("lead_time_days_avg",7)))
                        col_=C_RED if dos<lt*0.3 else (C_AMBER if dos<lt*0.7 else C_GREEN)
                        fg=go.Figure(go.Indicator(mode="gauge+number",value=round(dos,1),
                            number={"suffix":" days","font":{"color":col_,"size":16}},
                            gauge={"axis":{"range":[0,max(lt*1.5,1)],"tickcolor":C_GRAY},
                                   "bar":{"color":col_},"bgcolor":C_NAVY,"bordercolor":C_BORDER,
                                   "threshold":{"line":{"color":C_RED,"width":3},"value":lt}},
                            title={"text":f"{str(r.get('sku_id',''))[:8]}<br>{str(r.get('product_name',''))[:14]}",
                                   "font":{"color":C_WHITE,"size":9}}))
                        fg.update_layout(paper_bgcolor="rgba(0,0,0,0)",height=160,margin=dict(l=8,r=8,t=28,b=8))
                        PC(fg,f"gauge_{i}")
                st.markdown("</div>", unsafe_allow_html=True)

        # Alert table
        st.markdown('<div class="card">', unsafe_allow_html=True)
        sh("Live Alert Table",f"Sorted by {sort_by} — {len(s)} alerts")
        al_cols=[c for c in ["sku_id","product_name","category","alert_tier",
                              "days_of_supply_current","units_on_hand","total_available",
                              "reorder_point_optimized","safety_stock_seasonal",
                              "units_to_order","revenue_at_risk","priority_score",
                              "supplier_name","is_perishable","is_phantom","is_on_promo"] if c in s.columns]
        st.dataframe(s[al_cols].head(500),use_container_width=True,height=360,hide_index=True)
        dl_data=s[al_cols].to_csv(index=False)
        st.download_button(f"Download Morning Report — {store_id}",data=dl_data,
            file_name=f"{store_id}_morning_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",key="store_dl")
        st.markdown("</div>", unsafe_allow_html=True)

        # Store AI
        st.markdown("<hr>", unsafe_allow_html=True)
        sh("AI Assistant — Store Intelligence","")
        try:
            import importlib, sys as _sys5
            _sp5=str(Path(__file__).resolve().parent)
            if _sp5 not in _sys5.path: _sys5.path.insert(0,_sp5)
            import store_ai as _stai
            importlib.reload(_stai)
            _stai_ok=True
        except:
            _stai_ok=False
        _st_key="store"
        if _st_key not in st.session_state["chat"]:
            st.session_state["chat"][_st_key]=[]
        _st_msgs=st.session_state["chat"][_st_key]
        _st_quick=[
            ("Order now",f"What should I order right now at store {store_id}?"),
            ("Daily loss","What is my daily revenue loss if I do nothing?"),
            ("Supplier call","Which supplier should I call first?"),
            ("Phantom check","Which phantom SKUs need aisle check?"),
            ("Weekend risk","Which products stock out before this weekend?"),
        ]
        _stcols=st.columns(len(_st_quick))
        for _sti,(_stl,_stq) in enumerate(_st_quick):
            with _stcols[_sti]:
                if st.button(_stl,key=f"qq_st_{_sti}",use_container_width=True):
                    st.session_state["_pend_st"]=_stq
        for _m in _st_msgs:
            if _m["role"]=="user":
                _uh='<div class="chat-label" style="text-align:right">You</div>'
                _uh+='<div class="chat-user">'+str(_m["content"])+'</div>'
                st.markdown(_uh,unsafe_allow_html=True)
            else:
                _stxt=str(_m["content"]).replace("<","&lt;").replace(">","&gt;").replace(chr(10),"<br>")
                _ah='<div class="chat-label">AI</div><div class="chat-ai">'+_stxt+'</div>'
                st.markdown(_ah,unsafe_allow_html=True)
                if _m.get("fig") is not None:
                    try:
                        st.plotly_chart(_m["fig"],use_container_width=True,
                            config={"displayModeBar":False},key="st_chart_"+str(_st_msgs.index(_m)))
                    except: pass
        if not _st_msgs:
            st.markdown(f'<div class="chat-ai">Ask about what to order now, daily revenue loss, which supplier to call, phantom inventory, or weekend stockout risk at store {store_id}.</div>',unsafe_allow_html=True)
        _st_pend=st.session_state.pop("_pend_st",None)
        _st_in=st.chat_input(f"Ask about store {store_id}...",key="ci_store")
        _st_in=_st_in or _st_pend
        if _st_in:
            _st_msgs.append({"role":"user","content":_st_in,"fig":None})
            with st.spinner("Analysing store data..."):
                if _stai_ok:
                    _stxt2,_sfig2=_stai.answer(_st_in,na,store_id)
                else:
                    _stxt2="store_ai module not loaded."
                    _sfig2=None
            _st_msgs.append({"role":"assistant","content":_stxt2,"fig":_sfig2})
            st.rerun()
    else:
        st.markdown(f'<div style="color:{C_GRAY};padding:40px;text-align:center">Run python src/analytics.py first.</div>',unsafe_allow_html=True)

elif PAGE == "What-If Simulator":
    sh("What-If Simulator","Safety stock scenarios · supplier delay impact · 3-service-level comparison")
    st.markdown("<hr>", unsafe_allow_html=True)
    repl = D["repl"]

    # ── Store + SKU selectors ─────────────────────────────────
    wf1,wf2,wf3 = st.columns([2,2,1.5])
    with wf1:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store</div>', unsafe_allow_html=True)
        wf_store = st.selectbox("wf_s",[S_NAMES.get(s,s) for s in STORES[:80]],label_visibility="collapsed",key="wf_store")
        wf_sid = wf_store.split(" — ")[0] if wf_store else None
    with wf2:
        wf_skus = sorted(repl[repl["store_id"]==wf_sid]["sku_id"].dropna().unique().tolist()) if not repl.empty and wf_sid and "store_id" in repl.columns else []
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">SKU</div>', unsafe_allow_html=True)
        wf_sku = st.selectbox("wf_sku",wf_skus,label_visibility="collapsed",key="wf_sku") if wf_skus else None
    with wf3:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Apply to tier</div>', unsafe_allow_html=True)
        wf_tier = st.selectbox("wf_tier",["This SKU","All Premium","All High","All Low","Entire Store"],label_visibility="collapsed",key="wf_tier")
    st.markdown("<hr>", unsafe_allow_html=True)

    # Get base SKU data
    if not repl.empty and wf_sid and wf_sku:
        sku_data = repl[(repl["store_id"]==wf_sid)&(repl["sku_id"]==wf_sku)]
        if sku_data.empty: sku_data = repl[repl["store_id"]==wf_sid].head(1)
        base = sku_data.iloc[0] if not sku_data.empty else None
    else:
        base = repl.iloc[0] if not repl.empty else None

    if base is not None:
        from scipy import stats as _wsc
        import numpy as _wnp
        mu   = float(base.get("mu_daily",10))
        sig  = float(base.get("sigma_daily",3))
        lt   = float(base.get("lead_time_days_avg",7))
        lt_s = float(base.get("lead_time_days_std",1))
        rel  = float(base.get("reliability_score",0.9))

        # ── Sliders ──────────────────────────────────────────────
        sh("Adjust Parameters")
        sl1,sl2,sl3 = st.columns(3)
        with sl1:
            z_val  = st.slider("Z score (service level)",1.0,3.0,1.96,0.01,key="wf_z",
                help="1.64=95% · 1.96=97.5% · 2.33=99%")
            demand_mult = st.slider("Demand multiplier",0.5,2.0,1.0,0.05,key="wf_dm",
                help="1.0 = baseline demand, 1.2 = 20% demand surge")
        with sl2:
            lt_extra = st.slider("Supplier delay (extra days)",0,14,0,1,key="wf_lt",
                help="Simulate supplier delivering N days late")
            rel_adj  = st.slider("Supplier reliability",0.5,1.0,rel,0.01,key="wf_rel",
                help="Current reliability: "+str(round(rel,2)))
        with sl3:
            unit_cost = st.slider("Unit cost ($)",1,200,30,1,key="wf_uc")
            holding_days = st.slider("Holding period (days)",7,90,30,1,key="wf_hd")

        # ── Live formula ─────────────────────────────────────────
        eff_mu  = mu * demand_mult
        eff_lt  = lt + lt_extra
        eff_sig = sig * demand_mult
        ss_calc = z_val * eff_sig * _wnp.sqrt(eff_lt)
        rop_calc = eff_mu * eff_lt + ss_calc
        carry   = ss_calc * unit_cost * holding_days / 30
        sl_pct  = float(_wsc.norm.cdf(z_val)) * 100

        st.markdown("<hr>", unsafe_allow_html=True)
        sh("Live Formula Result")
        r1,r2,r3,r4,r5 = st.columns(5)
        with r1: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_TEAL};border-radius:8px;padding:12px;text-align:center"><div style="font-size:22px;font-weight:800;color:{C_TEAL}">{ss_calc:.1f}u</div><div style="font-size:10px;color:{C_GRAY}">Safety Stock</div><div style="font-size:9px;color:{C_GRAY};margin-top:2px">Z × σ × √LT</div></div>', unsafe_allow_html=True)
        with r2: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_AMBER};border-radius:8px;padding:12px;text-align:center"><div style="font-size:22px;font-weight:800;color:{C_AMBER}">{rop_calc:.1f}u</div><div style="font-size:10px;color:{C_GRAY}">Reorder Point</div><div style="font-size:9px;color:{C_GRAY};margin-top:2px">μ × LT + SS</div></div>', unsafe_allow_html=True)
        with r3:
            sl_col = C_GREEN if sl_pct>=97 else (C_AMBER if sl_pct>=94 else C_RED)
            st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {sl_col};border-radius:8px;padding:12px;text-align:center"><div style="font-size:22px;font-weight:800;color:{sl_col}">{sl_pct:.1f}%</div><div style="font-size:10px;color:{C_GRAY}">Service Level</div><div style="font-size:9px;color:{C_GRAY};margin-top:2px">norm.cdf(Z={z_val:.2f})</div></div>', unsafe_allow_html=True)
        with r4: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_PURPLE};border-radius:8px;padding:12px;text-align:center"><div style="font-size:22px;font-weight:800;color:{C_PURPLE}">{eff_lt:.1f}d</div><div style="font-size:10px;color:{C_GRAY}">Effective Lead Time</div><div style="font-size:9px;color:{C_GRAY};margin-top:2px">{lt:.0f} + {lt_extra}d delay</div></div>', unsafe_allow_html=True)
        with r5:
            cc_col = C_RED if carry > unit_cost*30 else C_AMBER
            st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {cc_col};border-radius:8px;padding:12px;text-align:center"><div style="font-size:22px;font-weight:800;color:{cc_col}">${carry:,.0f}</div><div style="font-size:10px;color:{C_GRAY}">Carrying Cost</div><div style="font-size:9px;color:{C_GRAY};margin-top:2px">SS × cost × {holding_days}d</div></div>', unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)

        # ── 3-Scenario comparison ─────────────────────────────────
        sh("3-Scenario Comparison")
        scenarios = [
            ("95% SL (Conservative)", 1.645),
            ("97.5% SL (Recommended)", 1.960),
            ("99% SL (Aggressive)", 2.326),
        ]
        import pandas as _pdwf
        sc_rows = []
        for label, z in scenarios:
            ss_ = z * eff_sig * _wnp.sqrt(eff_lt)
            rop_ = eff_mu * eff_lt + ss_
            carry_ = ss_ * unit_cost * holding_days / 30
            sl_ = float(_wsc.norm.cdf(z)) * 100
            sc_rows.append({"Scenario":label,"Z Score":f"{z:.3f}",
                "Safety Stock":f"{ss_:.1f}u","Reorder Point":f"{rop_:.1f}u",
                "Service Level":f"{sl_:.2f}%","Carrying Cost":f"${carry_:,.0f}"})
        st.dataframe(_pdwf.DataFrame(sc_rows),use_container_width=True,height=145,hide_index=True)

        # Chart
        zs = [s[1] for s in scenarios]
        labels_sc = [s[0].split(" ")[0] for s in scenarios]
        ss_vals  = [z*eff_sig*_wnp.sqrt(eff_lt) for z in zs]
        rop_vals = [eff_mu*eff_lt+ss for ss in ss_vals]
        carry_vals=[ss*unit_cost*holding_days/30 for ss in ss_vals]
        fig_wf = go.Figure()
        fig_wf.add_trace(go.Bar(name="Safety Stock",x=labels_sc,y=ss_vals,
            marker_color=C_TEAL,text=[f"{v:.0f}u" for v in ss_vals],
            textposition="outside",textfont=dict(color=C_WHITE,size=10)))
        fig_wf.add_trace(go.Bar(name="Reorder Point",x=labels_sc,y=rop_vals,
            marker_color=C_AMBER,text=[f"{v:.0f}u" for v in rop_vals],
            textposition="outside",textfont=dict(color=C_WHITE,size=10)))
        fig_wf.update_layout(**DL("Safety Stock · ROP by scenario",height=260),barmode="group")
        PC(fig_wf,"wf_sc")

        # Sensitivity sweep chart
        st.markdown("<hr>", unsafe_allow_html=True)
        sh("Sensitivity Sweep","How does safety stock change as Z score varies?")
        z_range = [round(1.0+i*0.1,1) for i in range(16)]  # 1.0 to 2.5
        ss_range = [z*eff_sig*_wnp.sqrt(eff_lt) for z in z_range]
        sl_range = [float(_wsc.norm.cdf(z))*100 for z in z_range]
        fig_sw = go.Figure()
        fig_sw.add_trace(go.Scatter(x=z_range,y=ss_range,name="Safety Stock (units)",
            line=dict(color=C_TEAL,width=2),mode="lines+markers",
            marker=dict(size=6,color=C_TEAL)))
        fig_sw.add_trace(go.Scatter(x=z_range,y=sl_range,name="Service Level %",
            line=dict(color=C_AMBER,width=2,dash="dot"),mode="lines",
            yaxis="y2"))
        fig_sw.add_vline(x=z_val,line_dash="dash",line_color=C_RED,
            annotation_text=f"Current Z={z_val:.2f}",
            annotation_font=dict(color=C_RED,size=9))
        fig_sw.update_layout(**DL("Safety Stock vs Z Score sensitivity",height=260),
            yaxis2=dict(title="Service Level %",overlaying="y",side="right",
                showgrid=False,tickfont=dict(color=C_AMBER),range=[90,100]))
        fig_sw.update_xaxes(title="Z Score")
        PC(fig_sw,"wf_sweep")

    else:
        st.markdown(f'<div style="color:{C_GRAY};padding:40px;text-align:center">Select a store and SKU above to start simulation.</div>', unsafe_allow_html=True)

    # What-If AI
    st.markdown("<hr>", unsafe_allow_html=True)
    sh("AI Assistant — What-If Intelligence")
    try:
        import importlib, sys as _sys9
        _sp9=str(Path(__file__).resolve().parent)
        if _sp9 not in _sys9.path: _sys9.path.insert(0,_sp9)
        import whatif_ai as _wai
        importlib.reload(_wai)
        _wai_ok=True
    except:
        _wai_ok=False
    _wf_key="whatif"
    if _wf_key not in st.session_state["chat"]: st.session_state["chat"][_wf_key]=[]
    _wf_msgs=st.session_state["chat"][_wf_key]
    _wf_quick=[
        ("Explain Z score","What does the Z score mean for my safety stock?"),
        ("Supplier delay","What happens if my supplier gets 7 days slower?"),
        ("Compare scenarios","Show me 95% vs 97.5% vs 99% service level trade-offs"),
        ("Apply to Premium","What does applying these settings to all Premium stores cost?"),
    ]
    _wfcols=st.columns(len(_wf_quick))
    for _wfi,(_wfl,_wfq) in enumerate(_wf_quick):
        with _wfcols[_wfi]:
            if st.button(_wfl,key=f"qq_wf_{_wfi}",use_container_width=True):
                st.session_state["_pend_wf"]=_wfq
    for _m in _wf_msgs:
        if _m["role"]=="user":
            _uh='<div class="chat-label" style="text-align:right">You</div>'
            _uh+='<div class="chat-user">'+str(_m["content"])+'</div>'
            st.markdown(_uh,unsafe_allow_html=True)
        else:
            _wtxt=str(_m["content"]).replace("<","&lt;").replace(">","&gt;").replace(chr(10),"<br>")
            _ah='<div class="chat-label">AI</div><div class="chat-ai">'+_wtxt+'</div>'
            st.markdown(_ah,unsafe_allow_html=True)
    if not _wf_msgs:
        st.markdown('<div class="chat-ai">Adjust the sliders above then ask: explain Z score · supplier delay impact · compare 95 vs 99% · apply to Premium stores.</div>',unsafe_allow_html=True)
    _wf_pend=st.session_state.pop("_pend_wf",None)
    _wf_in=st.chat_input("Ask about this simulation...",key="ci_whatif")
    _wf_in=_wf_in or _wf_pend
    if _wf_in:
        _wf_msgs.append({"role":"user","content":_wf_in})
        with st.spinner("Analysing simulation..."):
            if _wai_ok and base is not None:
                _wtxt2=_wai.answer(_wf_in,mu,sig,lt,lt_extra,z_val,unit_cost,holding_days,rel,
                    wf_sku or "",wf_sid or "")
            else:
                _wtxt2="Select a store and SKU above to activate the simulation."
        _wf_msgs.append({"role":"assistant","content":_wtxt2})
        st.rerun()

elif PAGE == "Store Comparison":
    sh("Store Comparison","Root cause · 5-factor breakdown · supplier scorecard · phantom analysis")
    st.markdown("<hr>", unsafe_allow_html=True)

    # Store selectors
    store_opts_cmp = [S_NAMES.get(s,s) for s in STORES[:80]]
    cmp1,cmp2,cmp3,cmp4 = st.columns([2,2,2,1.5])
    with cmp1:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store A</div>', unsafe_allow_html=True)
        cmp_sa = st.selectbox("cmp_sa",store_opts_cmp,label_visibility="collapsed",key="cmp_sa")
        cmp_sa_id = cmp_sa.split(" — ")[0] if cmp_sa else STORES[0]
    with cmp2:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store B</div>', unsafe_allow_html=True)
        cmp_sb = st.selectbox("cmp_sb",store_opts_cmp,index=min(1,len(store_opts_cmp)-1),label_visibility="collapsed",key="cmp_sb")
        cmp_sb_id = cmp_sb.split(" — ")[0] if cmp_sb else STORES[1]
    with cmp3:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store C (optional)</div>', unsafe_allow_html=True)
        cmp_sc = st.selectbox("cmp_sc",["None"]+store_opts_cmp,label_visibility="collapsed",key="cmp_sc")
        cmp_sc_id = cmp_sc.split(" — ")[0] if cmp_sc and cmp_sc!="None" else None
    with cmp4:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Category</div>', unsafe_allow_html=True)
        cmp_cats = sorted(na["category"].dropna().unique().tolist()) if not na.empty and "category" in na.columns else []
        cmp_cat = st.selectbox("cmp_cat",["All"]+cmp_cats,label_visibility="collapsed",key="cmp_cat")
    st.markdown("<hr>", unsafe_allow_html=True)

    # Get store data
    def get_store_alerts(sid):
        s = na[na["store_id"]==sid].copy() if not na.empty and sid else pd.DataFrame()
        if cmp_cat != "All" and not s.empty and "category" in s.columns: s=s[s["category"]==cmp_cat]
        return s

    sa_alerts = get_store_alerts(cmp_sa_id)
    sb_alerts = get_store_alerts(cmp_sb_id)
    sc_alerts = get_store_alerts(cmp_sc_id) if cmp_sc_id else pd.DataFrame()

    def kpi_for(alerts, sid):
        ns_r = ns[ns["store_id"]==sid] if not ns.empty and sid in ns["store_id"].values else pd.DataFrame()
        return {
            "critical": int((alerts["alert_tier"]=="CRITICAL").sum()) if not alerts.empty and "alert_tier" in alerts.columns else 0,
            "warning":  int((alerts["alert_tier"]=="WARNING").sum()) if not alerts.empty and "alert_tier" in alerts.columns else 0,
            "rar": float(alerts["revenue_at_risk"].sum()) if not alerts.empty and "revenue_at_risk" in alerts.columns else 0,
            "phantom": int(alerts["is_phantom"].sum()) if not alerts.empty and "is_phantom" in alerts.columns else 0,
            "uto": float(alerts["units_to_order"].sum()) if not alerts.empty and "units_to_order" in alerts.columns else 0,
            "dos": float(alerts["days_of_supply_current"].mean()) if not alerts.empty and "days_of_supply_current" in alerts.columns else 0,
            "name": ns_r["store_name"].iloc[0] if not ns_r.empty and "store_name" in ns_r.columns else sid,
            "format": ns_r["store_format"].iloc[0] if not ns_r.empty and "store_format" in ns_r.columns else "",
            "urgency": float(ns_r["urgency_score"].iloc[0]) if not ns_r.empty and "urgency_score" in ns_r.columns else 0,
        }

    ksa = kpi_for(sa_alerts, cmp_sa_id)
    ksb = kpi_for(sb_alerts, cmp_sb_id)
    ksc = kpi_for(sc_alerts, cmp_sc_id) if cmp_sc_id else None

    # KPI comparison cards
    n_stores = 3 if cmp_sc_id else 2
    store_list = [(cmp_sa_id, ksa, C_TEAL), (cmp_sb_id, ksb, C_AMBER)]
    if cmp_sc_id and ksc: store_list.append((cmp_sc_id, ksc, C_GREEN))

    card_cols = st.columns(n_stores)
    for ci,(sid,k,col) in enumerate(store_list):
        with card_cols[ci]:
            rar_k = k["rar"]/1e3
            st.markdown(f'<div style="background:{C_CARD};border:2px solid {col}44;border-top:4px solid {col};border-radius:10px;padding:16px"><div style="font-size:13px;font-weight:800;color:{col};margin-bottom:10px">{sid}</div><div style="font-size:11px;color:{C_GRAY}">{k["name"]}</div><div style="margin-top:10px"><div style="display:flex;justify-content:space-between;margin-bottom:4px"><span style="color:{C_GRAY};font-size:10px">Revenue at Risk</span><span style="color:{C_RED};font-weight:700">${rar_k:.0f}K</span></div><div style="display:flex;justify-content:space-between;margin-bottom:4px"><span style="color:{C_GRAY};font-size:10px">CRITICAL SKUs</span><span style="color:{C_RED};font-weight:700">{k["critical"]}</span></div><div style="display:flex;justify-content:space-between;margin-bottom:4px"><span style="color:{C_GRAY};font-size:10px">Warning SKUs</span><span style="color:{C_AMBER};font-weight:700">{k["warning"]}</span></div><div style="display:flex;justify-content:space-between;margin-bottom:4px"><span style="color:{C_GRAY};font-size:10px">Phantom SKUs</span><span style="color:{C_PURPLE};font-weight:700">{k["phantom"]}</span></div><div style="display:flex;justify-content:space-between;margin-bottom:4px"><span style="color:{C_GRAY};font-size:10px">Avg Days Supply</span><span style="color:{C_WHITE};font-weight:700">{k["dos"]:.1f}d</span></div><div style="display:flex;justify-content:space-between"><span style="color:{C_GRAY};font-size:10px">Urgency Score</span><span style="color:{C_WHITE};font-weight:700">{k["urgency"]:.0f}</span></div></div></div>', unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)

    # 5-Factor comparison chart
    sh("5-Factor Priority Score Breakdown","F1=Tier · F2=Revenue · F3=DOS · F4=Supplier · F5=Seasonal")
    cats_f = ["F1: Alert Tier","F2: Revenue/SKU","F3: Days Supply","F4: Supplier Risk","F5: Seasonal"]
    fig_5f = go.Figure()
    for sid,alerts,col in [(cmp_sa_id,sa_alerts,C_TEAL),(cmp_sb_id,sb_alerts,C_AMBER)]+\
        ([(cmp_sc_id,sc_alerts,C_GREEN)] if cmp_sc_id else []):
        if not alerts.empty:
            vals = [
                float(alerts["f1_tier"].mean()) if "f1_tier" in alerts.columns else 0,
                float(alerts["f2_revenue"].mean()) if "f2_revenue" in alerts.columns else 0,
                float(alerts["f3_dos"].mean()) if "f3_dos" in alerts.columns else 0,
                float(alerts["f4_supplier"].mean()) if "f4_supplier" in alerts.columns else 0,
                abs(float(alerts["f5_seasonal"].mean())) if "f5_seasonal" in alerts.columns else 0,
            ]
            fig_5f.add_trace(go.Bar(name=sid, x=cats_f, y=vals,
                marker_color=col,
                text=[f"{v:.2f}" for v in vals],
                textposition="outside",textfont=dict(color=C_WHITE,size=8)))
    fig_5f.update_layout(**DL("",height=280),barmode="group")
    PC(fig_5f,"cmp_5f")

    # Alert tier comparison
    c1_,c2_ = st.columns(2)
    with c1_:
        sh("Revenue at Risk Comparison")
        fig_rar = go.Figure()
        for sid,k,col in store_list:
            by_cat = get_store_alerts(sid)
            if not by_cat.empty and "category" in by_cat.columns and "revenue_at_risk" in by_cat.columns:
                cr = by_cat.groupby("category")["revenue_at_risk"].sum().nlargest(8)
                fig_rar.add_trace(go.Bar(name=sid, x=list(cr.index), y=cr.values,
                    marker_color=col,
                    text=[f"${v:,.0f}" for v in cr.values],
                    textposition="outside",textfont=dict(color=C_WHITE,size=7)))
        fig_rar.update_layout(**DL("",height=300),barmode="group")
        fig_rar.update_xaxes(tickangle=-30)
        PC(fig_rar,"cmp_rar")
    with c2_:
        sh("Phantom SKUs Breakdown")
        fig_ph = go.Figure()
        for sid,alerts,col in [(cmp_sa_id,sa_alerts,C_TEAL),(cmp_sb_id,sb_alerts,C_AMBER)]+\
            ([(cmp_sc_id,sc_alerts,C_GREEN)] if cmp_sc_id else []):
            if not alerts.empty and "is_phantom" in alerts.columns and "category" in alerts.columns:
                ph_c = alerts[alerts["is_phantom"]==True].groupby("category").size()
                if not ph_c.empty:
                    fig_ph.add_trace(go.Bar(name=sid, x=list(ph_c.index), y=ph_c.values,
                        marker_color=col,
                        text=ph_c.values,textposition="outside",
                        textfont=dict(color=C_WHITE,size=8)))
        fig_ph.update_layout(**DL("",height=300),barmode="group")
        fig_ph.update_xaxes(tickangle=-30)
        PC(fig_ph,"cmp_ph")

    # Supplier 4-factor comparison table
    sh("Supplier 4-Factor Scorecard Comparison")
    sup_rows = []
    for sid,alerts in [(cmp_sa_id,sa_alerts),(cmp_sb_id,sb_alerts)]+\
        ([(cmp_sc_id,sc_alerts)] if cmp_sc_id else []):
        if not alerts.empty and "supplier_name" in alerts.columns:
            sg = alerts.groupby("supplier_name").agg(
                critical=("alert_tier",lambda x:(x=="CRITICAL").sum()),
                revenue=("revenue_at_risk","sum"),
                f4=("f4_supplier","mean")
            ).reset_index().nlargest(3,"revenue")
            for _,r in sg.iterrows():
                _sup_sc=D["sup_sc"]; sc_row = _sup_sc[_sup_sc["supplier_name"]==r["supplier_name"]].iloc[0] if not _sup_sc.empty and "supplier_name" in _sup_sc.columns and r["supplier_name"] in _sup_sc["supplier_name"].values else pd.Series()
                sup_rows.append({
                    "Store":sid,"Supplier":r["supplier_name"],
                    "Critical SKUs":int(r["critical"]),"Revenue at Risk":f"${r['revenue']:,.0f}",
                    "Fill Rate %":f"{float(sc_row.get('avg_fulfillment_rate',0))*100:.1f}%" if hasattr(sc_row,"get") and sc_row.get("avg_fulfillment_rate") else "N/A",
                    "Late %":f"{float(sc_row.get('late_delivery_rate',0))*100:.1f}%" if hasattr(sc_row,"get") and sc_row.get("late_delivery_rate") else "N/A",
                    "Short %":f"{float(sc_row.get('short_delivery_rate',0))*100:.1f}%" if hasattr(sc_row,"get") and sc_row.get("short_delivery_rate") else "N/A",
                    "Risk Tier":str(sc_row.get("risk_tier","?")) if hasattr(sc_row,"get") else "?",
                })
    if sup_rows:
        st.dataframe(pd.DataFrame(sup_rows),use_container_width=True,height=280,hide_index=True)

    # Store Comparison AI
    st.markdown("<hr>", unsafe_allow_html=True)
    sh("AI Assistant — Store Comparison Intelligence","")
    try:
        import importlib, sys as _sys8
        _sp8=str(Path(__file__).resolve().parent)
        if _sp8 not in _sys8.path: _sys8.path.insert(0,_sp8)
        import store_comparison_ai as _scai
        importlib.reload(_scai)
        _scai_ok=True
    except Exception as _e:
        _scai_ok=False
    _cmp_key="comparison"
    if _cmp_key not in st.session_state["chat"]: st.session_state["chat"][_cmp_key]=[]
    _cmp_msgs=st.session_state["chat"][_cmp_key]
    _cmp_quick=[
        ("Why is A worse?",f"Why is {cmp_sa_id} worse than {cmp_sb_id}? Show root cause breakdown"),
        ("Which to visit first?",f"Which store should I visit first: {cmp_sa_id} or {cmp_sb_id}?"),
        ("3-store compare","Compare all three stores side by side with key metrics"),
        ("Supplier compare",f"Compare supplier performance between {cmp_sa_id} and {cmp_sb_id}"),
    ]
    _cmpcols=st.columns(len(_cmp_quick))
    for _cmpi,(_cmpl,_cmpq) in enumerate(_cmp_quick):
        with _cmpcols[_cmpi]:
            if st.button(_cmpl,key=f"qq_cmp_{_cmpi}",use_container_width=True):
                st.session_state["_pend_cmp"]=_cmpq
    for _m in _cmp_msgs:
        if _m["role"]=="user":
            _uh='<div class="chat-label" style="text-align:right">You</div>'
            _uh+='<div class="chat-user">'+str(_m["content"])+'</div>'
            st.markdown(_uh,unsafe_allow_html=True)
        else:
            _ctxt=str(_m["content"]).replace("<","&lt;").replace(">","&gt;").replace(chr(10),"<br>")
            _ah='<div class="chat-label">AI</div><div class="chat-ai">'+_ctxt+'</div>'
            st.markdown(_ah,unsafe_allow_html=True)
            if _m.get("fig") is not None:
                try:
                    st.plotly_chart(_m["fig"],use_container_width=True,
                        config={"displayModeBar":False},key="cmp_chart_"+str(_cmp_msgs.index(_m)))
                except: pass
    if not _cmp_msgs:
        st.markdown('<div class="chat-ai">Select two or three stores above and ask: why is A worse · which to visit first · 3-store comparison · supplier performance comparison.</div>',unsafe_allow_html=True)
    _cmp_pend=st.session_state.pop("_pend_cmp",None)
    _cmp_in=st.chat_input(f"Ask about {cmp_sa_id} vs {cmp_sb_id}...",key="ci_comparison")
    _cmp_in=_cmp_in or _cmp_pend
    if _cmp_in:
        _cmp_msgs.append({"role":"user","content":_cmp_in,"fig":None})
        with st.spinner("Analysing store data..."):
            if _scai_ok:
                _ctxt2,_cfig2=_scai.answer(_cmp_in,na,ns,D["sup_sc"],cmp_sa_id,cmp_sb_id,cmp_sc_id)
            else:
                _ctxt2="store_comparison_ai module not loaded."
                _cfig2=None
        _cmp_msgs.append({"role":"assistant","content":_ctxt2,"fig":_cfig2})
        st.rerun()

elif PAGE == "Safety Stock & ROP":
    sh("Safety Stock & ROP","AI policy vs old policy · SS% slider · below-ROP list · 3 scenarios")
    st.markdown("<hr>", unsafe_allow_html=True)
    repl = D["repl"]
    sf1,sf2,sf3 = st.columns([2,1.5,1.5])
    with sf1:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store</div>', unsafe_allow_html=True)
        ss_store=st.selectbox("ss_s",[S_NAMES.get(s,s) for s in STORES[:80]],label_visibility="collapsed",key="ss_store")
        ss_sid=ss_store.split(" — ")[0] if ss_store else None
    with sf2:
        avail_cats_ss=sorted(repl["category"].dropna().unique().tolist()) if not repl.empty and "category" in repl.columns else []
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Category</div>', unsafe_allow_html=True)
        ss_cat=st.selectbox("ss_cat",["All"]+avail_cats_ss,label_visibility="collapsed",key="ss_cat")
    with sf3:
        avail_tiers_ss=sorted(repl["foot_traffic_tier"].dropna().unique().tolist()) if not repl.empty and "foot_traffic_tier" in repl.columns else []
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Traffic Tier</div>', unsafe_allow_html=True)
        ss_tier=st.selectbox("ss_tier",["All"]+avail_tiers_ss,label_visibility="collapsed",key="ss_tier")
    st.markdown("<hr>", unsafe_allow_html=True)
    if not repl.empty and ss_sid:
        sr_=repl[repl["store_id"]==ss_sid].copy() if "store_id" in repl.columns else repl.copy()
        if ss_cat!="All" and "category" in sr_.columns: sr_=sr_[sr_["category"]==ss_cat]
        if ss_tier!="All" and "foot_traffic_tier" in sr_.columns: sr_=sr_[sr_["foot_traffic_tier"]==ss_tier]

        # ── SS% slider ────────────────────────────────────────────
        sh("Safety Stock % Adjustment Simulator","Slide to simulate impact on SS, ROP, service level and carrying cost — pure Python")
        ss_adj=st.slider("Adjust safety stock by:",min_value=-30,max_value=30,value=0,step=5,format="%d%%",key="ss_adj")
        if "safety_stock_units" in sr_.columns and "mu_daily" in sr_.columns and "lead_time_days_avg" in sr_.columns:
            import numpy as _np; from scipy import stats as _sc
            sr_["ss_adjusted"]=(sr_["safety_stock_units"]*(1+ss_adj/100)).round(0)
            sr_["rop_adjusted"]=(sr_["mu_daily"]*sr_["lead_time_days_avg"]+sr_["ss_adjusted"]).round(0)
            if "sigma_daily" in sr_.columns:
                sr_["new_z"]=(sr_["ss_adjusted"]/(sr_["sigma_daily"]*_np.sqrt(sr_["lead_time_days_avg"])).clip(lower=0.01))
                sr_["new_sl"]=sr_["new_z"].apply(lambda z:_sc.norm.cdf(min(z,3.5))*100).round(1)
            else:
                sr_["new_sl"]=97.5
            avg_ss_old=sr_["safety_stock_units"].mean()
            avg_ss_new=sr_["ss_adjusted"].mean()
            avg_rop_new=sr_["rop_adjusted"].mean()
            avg_sl_new=sr_["new_sl"].mean()
            delta_carry=(avg_ss_new-avg_ss_old)*30
            ac1,ac2,ac3,ac4,ac5=st.columns(5)
            with ac1: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_GRAY};border-radius:8px;padding:12px;text-align:center"><div style="font-size:18px;font-weight:800;color:{C_WHITE}">{avg_ss_old:.0f}u</div><div style="font-size:10px;color:{C_GRAY}">Current avg SS</div></div>', unsafe_allow_html=True)
            with ac2:
                col_ss=C_GREEN if ss_adj>0 else (C_RED if ss_adj<0 else C_WHITE)
                st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {col_ss};border-radius:8px;padding:12px;text-align:center"><div style="font-size:18px;font-weight:800;color:{col_ss}">{avg_ss_new:.0f}u</div><div style="font-size:10px;color:{C_GRAY}">New SS ({ss_adj:+d}%)</div></div>', unsafe_allow_html=True)
            with ac3: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_AMBER};border-radius:8px;padding:12px;text-align:center"><div style="font-size:18px;font-weight:800;color:{C_AMBER}">{avg_rop_new:.0f}u</div><div style="font-size:10px;color:{C_GRAY}">New avg ROP</div></div>', unsafe_allow_html=True)
            with ac4:
                sl_col=C_GREEN if avg_sl_new>=97 else (C_AMBER if avg_sl_new>=94 else C_RED)
                st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {sl_col};border-radius:8px;padding:12px;text-align:center"><div style="font-size:18px;font-weight:800;color:{sl_col}">{avg_sl_new:.1f}%</div><div style="font-size:10px;color:{C_GRAY}">Est. service level</div></div>', unsafe_allow_html=True)
            with ac5:
                cc=C_RED if delta_carry>0 else C_GREEN
                st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {cc};border-radius:8px;padding:12px;text-align:center"><div style="font-size:18px;font-weight:800;color:{cc}">${delta_carry:+,.0f}</div><div style="font-size:10px;color:{C_GRAY}">Carrying cost Δ/SKU</div></div>', unsafe_allow_html=True)
        st.markdown("<hr>", unsafe_allow_html=True)

        # ── 3-Scenario comparison ─────────────────────────────────
        sh("3-Scenario Comparison","95% vs 97.5% vs 99% service level — pure Python, no AI")
        if "sigma_daily" in sr_.columns and "lead_time_days_avg" in sr_.columns and "mu_daily" in sr_.columns:
            from scipy import stats as _sc2; import numpy as _np2
            scenarios=[("95% SL",_sc2.norm.ppf(0.95),C_AMBER),("97.5% SL",_sc2.norm.ppf(0.975),C_TEAL),("99% SL",_sc2.norm.ppf(0.99),C_GREEN)]
            sc_results=[]
            for label,z,color in scenarios:
                ss_c=(z*sr_["sigma_daily"]*_np2.sqrt(sr_["lead_time_days_avg"])).mean()
                rop_c=(sr_["mu_daily"]*sr_["lead_time_days_avg"]+ss_c).mean()
                sc_results.append({"label":label,"z":z,"ss":ss_c,"rop":rop_c,"carry":ss_c*30,"color":color})
            import pandas as _pd_sc
            sc_df=_pd_sc.DataFrame([{"Scenario":r["label"],"Z Score":f"{r['z']:.4f}","Avg Safety Stock":f"{r['ss']:.1f}u","Avg Reorder Point":f"{r['rop']:.1f}u","Est. Carrying Cost/SKU":f"${r['carry']:,.0f}"} for r in sc_results])
            st.dataframe(sc_df,use_container_width=True,height=145,hide_index=True)
            fig_sc=go.Figure()
            for r in sc_results:
                fig_sc.add_trace(go.Bar(name=r["label"],x=["Safety Stock","Reorder Point"],y=[r["ss"],r["rop"]],marker_color=r["color"],
                    text=[f"{r['ss']:.0f}u",f"{r['rop']:.0f}u"],textposition="outside",textfont=dict(color=C_WHITE,size=9)))
            fig_sc.update_layout(**DL("SS and ROP by service level target",height=260),barmode="group")
            PC(fig_sc,"ss_scenarios")
        st.markdown("<hr>", unsafe_allow_html=True)

        # ── Below-ROP table ───────────────────────────────────────
        sh("SKUs Below Reorder Point","These products need ordering now — sorted by most urgent")
        if "reorder_point" in sr_.columns and "mu_daily" in sr_.columns and "target" in sr_.columns:
            sr_["units_short"]=(sr_["reorder_point"]-sr_["target"]).clip(lower=0)
            sr_["days_until_stockout"]=(sr_["target"]/sr_["mu_daily"].clip(lower=0.1)).round(1)
            below_rop=sr_[sr_["units_short"]>0].sort_values("days_until_stockout")
            if not below_rop.empty:
                st.markdown(f'<div style="font-size:13px;font-weight:700;color:{C_RED};margin-bottom:10px">{len(below_rop):,} SKUs below reorder point at {ss_sid}</div>', unsafe_allow_html=True)
                show_cols=[c for c in ["sku_id","category","foot_traffic_tier","mu_daily","sigma_daily",
                    "safety_stock_units","reorder_point","target","units_short","days_until_stockout","lead_time_days_avg"] if c in below_rop.columns]
                st.dataframe(below_rop[show_cols].head(100).rename(columns={
                    "mu_daily":"Avg Demand/day","sigma_daily":"Demand Std","safety_stock_units":"Safety Stock",
                    "reorder_point":"ROP","target":"In Stock","units_short":"Units Short",
                    "days_until_stockout":"Days Until OOS","lead_time_days_avg":"Lead Time Days",
                    "foot_traffic_tier":"Tier"}),use_container_width=True,height=300,hide_index=True)
            else:
                st.markdown(f'<div style="color:{C_GREEN};padding:16px;font-size:14px;font-weight:700">✓ All SKUs above reorder point — store is well stocked</div>', unsafe_allow_html=True)
        st.markdown("<hr>", unsafe_allow_html=True)

        # ── Charts ────────────────────────────────────────────────
        sc1_,sc2_=st.columns([1,2])
        with sc1_:
            sh("SS Status Distribution")
            if all(c in sr_.columns for c in ["safety_stock_units","reorder_point","target"]):
                us_=int((sr_["target"]<sr_["reorder_point"]-sr_["safety_stock_units"]).sum())
                ov_=int((sr_["target"]>sr_["reorder_point"]+sr_["safety_stock_units"]*2).sum())
                op_=len(sr_)-us_-ov_
                fig_ss=go.Figure(go.Pie(labels=["Understocked","Overstocked","Optimal"],
                    values=[us_,ov_,op_],hole=0.55,marker_colors=[C_RED,C_AMBER,C_GREEN],
                    textinfo="label+percent",textfont=dict(color=C_WHITE,size=10)))
                fig_ss.add_annotation(text=f"{us_+ov_+op_:,}<br>SKUs",x=0.5,y=0.5,
                    showarrow=False,font=dict(color=C_WHITE,size=12))
                fig_ss.update_layout(**DL("",height=260,showlegend=False))
                PC(fig_ss,"ss_donut")
            else: PC(EF(height=260),"ss_donut_e")
        with sc2_:
            sh("AI Policy vs Old Policy — by Traffic Tier")
            if all(c in sr_.columns for c in ["foot_traffic_tier","target_upto_old","target_upto_new"]):
                tc_=sr_.groupby("foot_traffic_tier").agg(old=("target_upto_old","mean"),new=("target_upto_new","mean")).reset_index()
                fig2_=go.Figure()
                fig2_.add_trace(go.Bar(x=tc_["foot_traffic_tier"],y=tc_["old"],name="Old policy",marker_color=C_RED))
                fig2_.add_trace(go.Bar(x=tc_["foot_traffic_tier"],y=tc_["new"],name="AI policy",marker_color=C_TEAL))
                fig2_.update_layout(**DL("Avg target units by tier",height=260,barmode="group"))
                PC(fig2_,"ss_compare")
            else: PC(EF(height=260),"ss_compare_e")

        # ── Full SKU table ────────────────────────────────────────
        sh("Full Safety Stock Table")
        tbl_c=[c for c in ["store_id","sku_id","category","foot_traffic_tier","mu_daily",
            "sigma_daily","safety_stock_units","reorder_point","target_upto_old","target_upto_new",
            "lead_time_days_avg","csl","reliability_score"] if c in sr_.columns]
        if tbl_c:
            st.dataframe(sr_[tbl_c].head(300).rename(columns={
                "mu_daily":"Avg Demand","sigma_daily":"Demand Std","safety_stock_units":"Safety Stock",
                "reorder_point":"ROP","target_upto_old":"Old Target","target_upto_new":"AI Target",
                "lead_time_days_avg":"Lead Time","csl":"Service Level","reliability_score":"Supplier Reliability"}),
                use_container_width=True,height=320,hide_index=True)

        # ── Safety Stock AI ───────────────────────────────────────
        st.markdown("<hr>", unsafe_allow_html=True)
        sh("AI Assistant — Safety Stock Intelligence","")
        try:
            import importlib, sys as _sys6
            _sp6=str(Path(__file__).resolve().parent)
            if _sp6 not in _sys6.path: _sys6.path.insert(0,_sp6)
            import ss_ai as _ssai
            importlib.reload(_ssai)
            _ssai_ok=True
        except:
            _ssai_ok=False
        _ss_key="ss"
        if _ss_key not in st.session_state["chat"]:
            st.session_state["chat"][_ss_key]=[]
        _ss_msgs=st.session_state["chat"][_ss_key]
        _ss_quick=[
            ("Understocked SKUs","Which SKUs are most understocked right now?"),
            ("Capital in overstock","How much capital is locked in overstock?"),
            ("Below ROP","Show me SKUs below reorder point with days until stockout"),
            ("Simulate -10%","What would happen if we reduced safety stock by 10%?"),
        ]
        _sscols=st.columns(len(_ss_quick))
        for _ssi,(_ssl,_ssq) in enumerate(_ss_quick):
            with _sscols[_ssi]:
                if st.button(_ssl,key=f"qq_ss_{_ssi}",use_container_width=True):
                    st.session_state["_pend_ss"]=_ssq
        for _m in _ss_msgs:
            if _m["role"]=="user":
                _uh='<div class="chat-label" style="text-align:right">You</div>'
                _uh+='<div class="chat-user">'+str(_m["content"])+'</div>'
                st.markdown(_uh,unsafe_allow_html=True)
            else:
                _stxt=str(_m["content"]).replace("<","&lt;").replace(">","&gt;").replace(chr(10),"<br>")
                _ah='<div class="chat-label">AI</div><div class="chat-ai">'+_stxt+'</div>'
                st.markdown(_ah,unsafe_allow_html=True)
                if _m.get("fig") is not None:
                    try:
                        st.plotly_chart(_m["fig"],use_container_width=True,
                            config={"displayModeBar":False},key="ss_chart_"+str(_ss_msgs.index(_m)))
                    except: pass
        if not _ss_msgs:
            st.markdown('<div class="chat-ai">Ask about understocked SKUs, capital locked in overstock, SKUs below ROP, or simulate a safety stock change.</div>',unsafe_allow_html=True)
        _ss_pend=st.session_state.pop("_pend_ss",None)
        _ss_in=st.chat_input("Ask about safety stock...",key="ci_ss")
        _ss_in=_ss_in or _ss_pend
        if _ss_in:
            _ss_msgs.append({"role":"user","content":_ss_in,"fig":None})
            with st.spinner("Analysing safety stock data..."):
                if _ssai_ok:
                    _stxt2,_sfig2=_ssai.answer(_ss_in,sr_,ss_sid)
                else:
                    _stxt2="ss_ai module not loaded."
                    _sfig2=None
            _ss_msgs.append({"role":"assistant","content":_stxt2,"fig":_sfig2})
            st.rerun()
    else:
        st.markdown(f'<div style="color:{C_GRAY};padding:40px;text-align:center">Select a store above.</div>', unsafe_allow_html=True)

elif PAGE == "Phantom Inventory":
    ph=D["phantom_c"]; ph_s=D["phantom_s"]
    sh("Phantom Inventory","Ghost stock detection · confidence scoring · aisle check list")
    st.markdown("<hr>", unsafe_allow_html=True)
    pf1,pf2,pf3 = st.columns(3)
    with pf1:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store</div>', unsafe_allow_html=True)
        ph_store_opts = ["All stores"] + sorted(ph["store_id"].dropna().unique().tolist()) if not ph.empty else ["All stores"]
        ph_store = st.selectbox("Store",ph_store_opts,label_visibility="collapsed",key="ph_store")
    with pf2:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Min Zero-Sales Days</div>', unsafe_allow_html=True)
        ph_min_days = st.selectbox("Days",[3,7,14,30],label_visibility="collapsed",key="ph_min_days")
    with pf3:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Confidence Filter</div>', unsafe_allow_html=True)
        ph_conf = st.selectbox("Conf",["All","High","Medium","Low"],label_visibility="collapsed",key="ph_conf")
    st.markdown("<hr>", unsafe_allow_html=True)
    ph_f = ph[ph["is_phantom_candidate"]==True].copy() if not ph.empty and "is_phantom_candidate" in ph.columns else ph.copy()
    if not ph_f.empty and "consec_zero_days" in ph_f.columns:
        ph_f = ph_f[ph_f["consec_zero_days"] >= ph_min_days]
    if ph_conf != "All" and not ph_f.empty and "phantom_confidence" in ph_f.columns:
        ph_f = ph_f[ph_f["phantom_confidence"]==ph_conf]
    ph_sid = ph_store if ph_store != "All stores" else None
    ph_f_store = ph_f[ph_f["store_id"]==ph_sid] if ph_sid and not ph_f.empty else ph_f
    total_ph = len(ph_f)
    high_ph  = int((ph_f["phantom_confidence"]=="High").sum()) if not ph_f.empty and "phantom_confidence" in ph_f.columns else 0
    med_ph   = int((ph_f["phantom_confidence"]=="Medium").sum()) if not ph_f.empty and "phantom_confidence" in ph_f.columns else 0
    low_ph   = int((ph_f["phantom_confidence"]=="Low").sum()) if not ph_f.empty and "phantom_confidence" in ph_f.columns else 0
    if not ph_f.empty:
        ph_f["daily_rev"] = (ph_f["rolling_14d_avg"].fillna(0)*ph_f["unit_price"].fillna(0))
        total_rev = float(ph_f["daily_rev"].sum())
    else:
        total_rev = 0
    kpi_row([("Phantom Candidates",f"{total_ph:,}",f"Min {ph_min_days} zero-sales days",C_PURPLE),
             ("High Confidence",f"{high_ph:,}","Immediate physical check",C_RED),
             ("Medium Confidence",f"{med_ph:,}","Monitor closely",C_AMBER),
             ("Low Confidence",f"{low_ph:,}","Watch",C_GRAY),
             ("Est. Daily Revenue at Risk",f"${total_rev:,.0f}","14d avg × unit price",C_AMBER)])
    rc1,rc2 = st.columns(2)
    with rc1:
        sh("Phantom by Store — Top 20")
        if not ph_s.empty and "phantom_count" in ph_s.columns:
            top20=ph_s.nlargest(20,"phantom_count")
            fig_=go.Figure(go.Bar(x=top20["phantom_count"],y=top20["store_id"],orientation="h",
                marker=dict(color=top20["phantom_count"],colorscale=[[0,C_CARD],[1,C_PURPLE]],showscale=False),
                text=top20["phantom_count"],textposition="outside",textfont=dict(color=C_WHITE,size=9)))
            fig_.update_layout(**DL("",height=380))
            PC(fig_,"ph_top20")
        else: PC(EF("Run analytics.py",380),"ph_top20_e")
    with rc2:
        sh("Phantom Age Distribution","How long have these been sitting?")
        if not ph_f.empty and "consec_zero_days" in ph_f.columns:
            import pandas as _pd2
            ph_f2 = ph_f.copy()
            ph_f2["age_band"] = _pd2.cut(ph_f2["consec_zero_days"],bins=[3,7,14,22,999],
                labels=["3-6 days","7-13 days","14-21 days","22+ days"],right=False)
            age_g = ph_f2.groupby("age_band",observed=True).agg(
                count=("sku_id","count"),
                high=("phantom_confidence",lambda x:(x=="High").sum())).reset_index()
            fig_a=go.Figure()
            fig_a.add_trace(go.Bar(x=age_g["age_band"],y=age_g["count"],name="All",
                marker_color=[C_PURPLE,C_AMBER,C_RED,C_RED],
                text=age_g["count"],textposition="outside",textfont=dict(color=C_WHITE,size=9)))
            fig_a.add_trace(go.Bar(x=age_g["age_band"],y=age_g["high"],name="High conf",
                marker_color="rgba(239,68,68,0.5)",
                text=age_g["high"],textposition="outside",textfont=dict(color=C_WHITE,size=9)))
            fig_a.update_layout(**DL("",height=380),barmode="group")
            PC(fig_a,"ph_age")
        else: PC(EF(height=380),"ph_age_e")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    sh("Phantom by Category","Stacked by confidence level")
    if not ph_f.empty and "category" in ph_f.columns and "phantom_confidence" in ph_f.columns:
        ph_cat=ph_f.groupby(["category","phantom_confidence"]).size().reset_index(name="count")
        conf_c={"High":C_RED,"Medium":C_AMBER,"Low":C_PURPLE}
        fig2_=px.bar(ph_cat,x="count",y="category",color="phantom_confidence",
            orientation="h",barmode="stack",color_discrete_map=conf_c)
        fig2_.update_layout(**DL("",height=360))
        PC(fig2_,"ph_cat")
    else: PC(EF(height=360),"ph_cat_e")
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown('<div class="card-glow-r">', unsafe_allow_html=True)
    sh("Aisle Check List","High confidence first — sorted by most days without sales")
    if not ph_f_store.empty:
        conf_order={"High":0,"Medium":1,"Low":2}
        ph_f_store2=ph_f_store.copy()
        ph_f_store2["conf_rank"]=ph_f_store2["phantom_confidence"].map(conf_order).fillna(3)
        ph_f_store2["daily_rev"]=ph_f_store2["rolling_14d_avg"].fillna(0)*ph_f_store2["unit_price"].fillna(0)
        aisle_cols=[c for c in ["store_id","sku_id","product_name","category",
                                 "phantom_confidence","consec_zero_days","daily_rev","store_name"] if c in ph_f_store2.columns]
        aisle_df=ph_f_store2.sort_values(["conf_rank","consec_zero_days"],ascending=[True,False])[aisle_cols].head(50)
        aisle_df=aisle_df.rename(columns={"sku_id":"SKU","product_name":"Product",
            "category":"Category","phantom_confidence":"Confidence",
            "consec_zero_days":"Zero-Sales Days","daily_rev":"Est $/day","store_name":"Store Name"})
        st.dataframe(aisle_df,use_container_width=True,height=340,hide_index=True)
        st.download_button("Download Aisle Check List CSV",
            data=aisle_df.to_csv(index=False),
            file_name=f"aisle_check_{ph_store}.csv",mime="text/csv",key="ph_dl")
    else:
        st.markdown(f'<div style="color:{C_GRAY};padding:20px">No phantom candidates with current filters.</div>', unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)
    sh("AI Assistant — Phantom Inventory Intelligence","")
    try:
        import importlib, sys as _sys3
        _sp3=str(Path(__file__).resolve().parent)
        if _sp3 not in _sys3.path: _sys3.path.insert(0,_sp3)
        import phantom_ai as _pai
        importlib.reload(_pai)
        _pai_ok=True
    except:
        _pai_ok=False
    _ph_key="phantom"
    if _ph_key not in st.session_state["chat"]:
        st.session_state["chat"][_ph_key]=[]
    _ph_msgs=st.session_state["chat"][_ph_key]
    _ph_quick=[
        ("Aisle check","Which products should I physically check first?"),
        ("Capital locked","How much capital is locked in phantom inventory?"),
        ("Worst category","Which category has the most ghost inventory?"),
        ("Age analysis","How long have these phantoms been sitting?"),
    ]
    _phcols=st.columns(len(_ph_quick))
    for _phi,(_phl,_phq) in enumerate(_ph_quick):
        with _phcols[_phi]:
            if st.button(_phl,key=f"qq_ph_{_phi}",use_container_width=True):
                st.session_state["_pend_ph"]=_phq
    for _m in _ph_msgs:
        if _m["role"]=="user":
            _uh='<div class="chat-label" style="text-align:right">You</div>'
            _uh+='<div class="chat-user">'+str(_m["content"])+'</div>'
            st.markdown(_uh,unsafe_allow_html=True)
        else:
            _ptxt=str(_m["content"]).replace("<","&lt;").replace(">","&gt;").replace(chr(10),"<br>")
            _ah='<div class="chat-label">AI</div><div class="chat-ai">'+_ptxt+'</div>'
            st.markdown(_ah,unsafe_allow_html=True)
            if _m.get("fig") is not None:
                try:
                    st.plotly_chart(_m["fig"],use_container_width=True,
                        config={"displayModeBar":False},key="ph_chart_"+str(_ph_msgs.index(_m)))
                except: pass
    if not _ph_msgs:
        st.markdown('<div class="chat-ai">Ask about aisle checks, capital locked in phantoms, worst category, or how old these phantoms are.</div>',unsafe_allow_html=True)
    _ph_pend=st.session_state.pop("_pend_ph",None)
    _ph_in=st.chat_input("Ask about phantom inventory...",key="ci_phantom")
    _ph_in=_ph_in or _ph_pend
    if _ph_in:
        _ph_msgs.append({"role":"user","content":_ph_in,"fig":None})
        with st.spinner("Analysing phantom data..."):
            if _pai_ok:
                _ptxt2,_pfig2=_pai.answer(_ph_in,ph,
                    store_id=ph_sid,min_days=ph_min_days,confidence_filter=ph_conf)
            else:
                _ptxt2="phantom_ai module not loaded."
                _pfig2=None
        _ph_msgs.append({"role":"assistant","content":_ptxt2,"fig":_pfig2})
        st.rerun()

elif PAGE == "Localization":
    ls=D["loc_scores"]; lc=D["loc_cat"]; ln=D["loc_net"]
    sh("Localization","Store vs store · category mismatch · slow movers · revenue uplift")
    st.markdown("<hr>", unsafe_allow_html=True)

    # KPI strip
    avg_score=float(ls["localization_score"].mean()) if not ls.empty and "localization_score" in ls.columns else 0
    total_mm=int((lc["mismatch_flag"].isin(["MISMATCH","EXTREME_MISMATCH"])).sum()) if not lc.empty and "mismatch_flag" in lc.columns else 0
    overperf=int((lc["mismatch_flag"]=="OVERPERFORMING").sum()) if not lc.empty and "mismatch_flag" in lc.columns else 0
    worst_store=ls.loc[ls["localization_score"].idxmin(),"store_id"] if not ls.empty else ""
    worst_score=float(ls["localization_score"].min()) if not ls.empty else 0
    up_col="total_rev_gap" if not ln.empty and "total_rev_gap" in ln.columns else None
    uplift=abs(float(ln[up_col].sum())) if up_col else 0
    kpi_row([
        ("Avg Score",f"{avg_score:.1f}","0=worst, 100=best",C_TEAL),
        ("Mismatched Combos",f"{total_mm:,}","Below region avg",C_RED),
        ("Overperforming",f"{overperf:,}","Above region avg",C_GREEN),
        ("Worst Store",worst_store,f"Score {worst_score:.1f}",C_AMBER),
        ("Revenue Gap",f"${uplift/1e6:.1f}M","If mismatches fixed",C_PURPLE),
    ])
    st.markdown("<hr>", unsafe_allow_html=True)

    # Store vs Store comparison
    sh("Store vs Store Comparison","How do two stores perform across the same categories?")
    store_opts=sorted(ls["store_id"].dropna().unique().tolist()) if not ls.empty else STORES[:80]
    name_map={}
    if not ls.empty and "store_name" in ls.columns:
        for _,r in ls.iterrows():
            name_map[r["store_id"]]=f"{r['store_id']} — {r.get('store_name','')}, {r.get('city','')}"
    else:
        name_map={s:S_NAMES.get(s,s) for s in store_opts}
    store_disp=[name_map.get(s,s) for s in store_opts]
    avail_cats=sorted(lc["category"].dropna().unique().tolist()) if not lc.empty and "category" in lc.columns else []

    lv1,lv2,lv3=st.columns([2,2,2])
    with lv1:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store A</div>', unsafe_allow_html=True)
        loc_sa=st.selectbox("loc_sa",store_disp,label_visibility="collapsed",key="loc_sa")
        sa_id=loc_sa.split(" — ")[0] if loc_sa else None
    with lv2:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store B</div>', unsafe_allow_html=True)
        loc_sb=st.selectbox("loc_sb",store_disp,index=min(1,len(store_disp)-1),label_visibility="collapsed",key="loc_sb")
        sb_id=loc_sb.split(" — ")[0] if loc_sb else None
    with lv3:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Category</div>', unsafe_allow_html=True)
        loc_cat=st.selectbox("loc_cat",["All categories"]+avail_cats,label_visibility="collapsed",key="loc_cat")

    if not lc.empty and sa_id and sb_id and sa_id!=sb_id:
        da=lc[lc["store_id"]==sa_id].copy()
        db=lc[lc["store_id"]==sb_id].copy()
        if loc_cat!="All categories": da=da[da["category"]==loc_cat]; db=db[db["category"]==loc_cat]
        if not da.empty and not db.empty:
            merged=da[["category","store_avg_daily_rev","rev_gap_pct","mismatch_flag"]].merge(
                db[["category","store_avg_daily_rev","rev_gap_pct","mismatch_flag"]],
                on="category",suffixes=(f"_{sa_id}",f"_{sb_id}"))
            fig_cmp=go.Figure()
            fig_cmp.add_trace(go.Bar(name=sa_id,x=merged["category"],
                y=merged[f"store_avg_daily_rev_{sa_id}"],marker_color=C_TEAL,
                text=[f"${v:.0f}/day" for v in merged[f"store_avg_daily_rev_{sa_id}"]],
                textposition="outside",textfont=dict(color=C_WHITE,size=8)))
            fig_cmp.add_trace(go.Bar(name=sb_id,x=merged["category"],
                y=merged[f"store_avg_daily_rev_{sb_id}"],marker_color=C_AMBER,
                text=[f"${v:.0f}/day" for v in merged[f"store_avg_daily_rev_{sb_id}"]],
                textposition="outside",textfont=dict(color=C_WHITE,size=8)))
            fig_cmp.update_layout(**DL(f"Daily revenue by category: {sa_id} vs {sb_id}",height=320),
                barmode="group")
            fig_cmp.update_xaxes(tickangle=-30)
            PC(fig_cmp,"loc_cmp_chart")
            # Table
            merged2=merged.copy()
            merged2[f"Rev/day {sa_id}"]=merged2[f"store_avg_daily_rev_{sa_id}"].round(0).apply(lambda x:f"${x:,.0f}")
            merged2[f"Rev/day {sb_id}"]=merged2[f"store_avg_daily_rev_{sb_id}"].round(0).apply(lambda x:f"${x:,.0f}")
            merged2[f"Gap% {sa_id}"]=merged2[f"rev_gap_pct_{sa_id}"].apply(lambda x:f"{x*100:+.1f}%")
            merged2[f"Gap% {sb_id}"]=merged2[f"rev_gap_pct_{sb_id}"].apply(lambda x:f"{x*100:+.1f}%")
            merged2[f"Status {sa_id}"]=merged2[f"mismatch_flag_{sa_id}"]
            merged2[f"Status {sb_id}"]=merged2[f"mismatch_flag_{sb_id}"]
            sc=["category",f"Rev/day {sa_id}",f"Gap% {sa_id}",f"Status {sa_id}",f"Rev/day {sb_id}",f"Gap% {sb_id}",f"Status {sb_id}"]
            st.dataframe(merged2[sc],use_container_width=True,height=280,hide_index=True)
            mc1,mc2=st.columns(2)
            sa_score=float(ls[ls["store_id"]==sa_id]["localization_score"].iloc[0]) if not ls[ls["store_id"]==sa_id].empty else 0
            sb_score=float(ls[ls["store_id"]==sb_id]["localization_score"].iloc[0]) if not ls[ls["store_id"]==sb_id].empty else 0
            sa_mm=int((da["mismatch_flag"].isin(["MISMATCH","EXTREME_MISMATCH"])).sum())
            sb_mm=int((db["mismatch_flag"].isin(["MISMATCH","EXTREME_MISMATCH"])).sum())
            with mc1:
                sc1c=C_GREEN if sa_score>sb_score else C_AMBER
                st.markdown(f'<div style="background:{C_CARD};border:1px solid {sc1c}44;border-radius:8px;padding:14px"><div style="color:{C_TEAL};font-size:11px;font-weight:700;margin-bottom:6px">{sa_id}</div><div style="font-size:22px;font-weight:800;color:{sc1c}">{sa_score:.1f}</div><div style="font-size:11px;color:{C_GRAY}">Localization score · {sa_mm} mismatches</div></div>', unsafe_allow_html=True)
            with mc2:
                sc2c=C_GREEN if sb_score>sa_score else C_AMBER
                st.markdown(f'<div style="background:{C_CARD};border:1px solid {sc2c}44;border-radius:8px;padding:14px"><div style="color:{C_AMBER};font-size:11px;font-weight:700;margin-bottom:6px">{sb_id}</div><div style="font-size:22px;font-weight:800;color:{sc2c}">{sb_score:.1f}</div><div style="font-size:11px;color:{C_GRAY}">Localization score · {sb_mm} mismatches</div></div>', unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)

    # Network charts
    lc1,lc2=st.columns(2)
    with lc1:
        sh("Top Mismatched Categories","Revenue gap across network")
        if not ln.empty and up_col:
            ln_s=ln.sort_values(up_col)
            fig_=go.Figure(go.Bar(x=abs(ln_s[up_col]),y=ln_s["category"],orientation="h",
                marker=dict(color=abs(ln_s[up_col]),colorscale=[[0,C_AMBER],[1,C_RED]],showscale=False),
                text=[f"${abs(v):,.0f}" for v in ln_s[up_col]],
                textposition="outside",textfont=dict(color=C_WHITE,size=9)))
            fig_.update_layout(**DL("",height=320))
            PC(fig_,"loc_net_cat")
        else: PC(EF(height=320),"loc_net_cat_e")
    with lc2:
        sh("Worst Stores by Score","Lower score = bigger opportunity")
        if not ls.empty and "localization_score" in ls.columns:
            worst=ls.nsmallest(15,"localization_score")
            fig2_=go.Figure(go.Bar(x=worst["localization_score"],
                y=worst["store_id"]+" "+worst.get("store_name","").fillna("").str[:10],
                orientation="h",
                marker=dict(color=worst["localization_score"],colorscale=[[0,C_RED],[1,C_AMBER]],showscale=False),
                text=[f"{v:.1f}" for v in worst["localization_score"]],
                textposition="outside",textfont=dict(color=C_WHITE,size=9)))
            fig2_.update_layout(**DL("",height=400))
            fig2_.update_yaxes(tickfont=dict(size=9),automargin=True)
            PC(fig2_,"loc_store")
        else: PC(EF(height=400),"loc_store_e")

    # Mismatch detail table
    sh("Category Mismatch Detail","Stores underperforming vs region average")
    if not lc.empty:
        show_lc=lc[lc["mismatch_flag"].isin(["MISMATCH","EXTREME_MISMATCH"])].copy() if "mismatch_flag" in lc.columns else lc.copy()
        show_lc["rev_gap_pct_fmt"]=(show_lc["rev_gap_pct"]*100).round(1).apply(lambda x:f"{x:+.1f}%")
        show_lc["store_avg_daily_rev"]=show_lc["store_avg_daily_rev"].round(0)
        show_lc["region_avg_daily_rev"]=show_lc["region_avg_daily_rev"].round(0)
        disp_cols=[c for c in ["store_id","region","category","store_avg_daily_rev","region_avg_daily_rev","rev_gap_pct_fmt","mismatch_flag","store_sku_count","z_score"] if c in show_lc.columns]
        st.dataframe(show_lc[disp_cols].sort_values("rev_gap_pct_fmt" if "rev_gap_pct_fmt" in show_lc.columns else disp_cols[0]).head(200).rename(columns={
            "store_id":"Store","region":"Region","category":"Category",
            "store_avg_daily_rev":"Store Rev/day","region_avg_daily_rev":"Region Avg/day",
            "rev_gap_pct_fmt":"Rev Gap %","mismatch_flag":"Status",
            "store_sku_count":"SKU Count","z_score":"Z Score"}),
            use_container_width=True,height=300,hide_index=True)
    st.markdown("<hr>", unsafe_allow_html=True)

    # Localization AI
    sh("AI Assistant — Localization Intelligence","")
    try:
        import importlib, sys as _sys7
        _sp7=str(Path(__file__).resolve().parent)
        if _sp7 not in _sys7.path: _sys7.path.insert(0,_sp7)
        import loc_ai as _lai
        importlib.reload(_lai)
        _lai_ok=True
    except:
        _lai_ok=False
    _lk="loc"
    if _lk not in st.session_state["chat"]: st.session_state["chat"][_lk]=[]
    _lmsgs=st.session_state["chat"][_lk]
    _lquick=[
        ("Biggest opportunity","Which store has the biggest revenue opportunity?"),
        ("Category to fix","Which category should I localize first?"),
        ("Compare stores",f"Compare {sa_id} vs {sb_id} {loc_cat}"),
        ("Slow movers","Which slow-moving SKUs are tying up capital?"),
    ]
    _lcols=st.columns(len(_lquick))
    for _li,(_ll,_lq) in enumerate(_lquick):
        with _lcols[_li]:
            if st.button(_ll,key=f"qq_loc_{_li}",use_container_width=True):
                st.session_state["_pend_loc"]=_lq
    for _m in _lmsgs:
        if _m["role"]=="user":
            _uh='<div class="chat-label" style="text-align:right">You</div>'
            _uh+='<div class="chat-user">'+str(_m["content"])+'</div>'
            st.markdown(_uh,unsafe_allow_html=True)
        else:
            _ltxt=str(_m["content"]).replace("<","&lt;").replace(">","&gt;").replace(chr(10),"<br>")
            _ah='<div class="chat-label">AI</div><div class="chat-ai">'+_ltxt+'</div>'
            st.markdown(_ah,unsafe_allow_html=True)
            if _m.get("fig") is not None:
                try:
                    st.plotly_chart(_m["fig"],use_container_width=True,
                        config={"displayModeBar":False},key="loc_chart_"+str(_lmsgs.index(_m)))
                except: pass
    if not _lmsgs:
        st.markdown('<div class="chat-ai">Ask about biggest revenue opportunity, which category to localize first, store comparison, or slow movers tying up capital.</div>',unsafe_allow_html=True)
    _lpend=st.session_state.pop("_pend_loc",None)
    _lin=st.chat_input("Ask about localization...",key="ci_loc")
    _lin=_lin or _lpend
    if _lin:
        _lmsgs.append({"role":"user","content":_lin,"fig":None})
        with st.spinner("Analysing localization data..."):
            if _lai_ok:
                _ltxt2,_lfig2=_lai.answer(_lin,ls,lc,ln,D["pred"],D["na"],sa_id,sb_id,loc_cat)
            else:
                _ltxt2="loc_ai module not loaded."
                _lfig2=None
        _lmsgs.append({"role":"assistant","content":_ltxt2,"fig":_lfig2})
        st.rerun()

elif PAGE == "Supplier Performance":
    sc=D["sup_sc"]
    if sc.empty and not D["suppliers"].empty: sc=D["suppliers"].copy()
    sh("Supplier Performance","Reliability · fill rate · stockout impact · dual-source analysis")
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Column detection ─────────────────────────────────────
    fill_col  = next((c for c in ["avg_fulfillment_rate","avg_fill_rate"] if c in sc.columns), None)
    late_col  = next((c for c in ["late_delivery_rate"] if c in sc.columns), None)
    short_col = next((c for c in ["short_delivery_rate"] if c in sc.columns), None)
    stock_col = next((c for c in ["stockout_events_caused","total_stockout_events"] if c in sc.columns), None)
    risk_col  = next((c for c in ["risk_tier"] if c in sc.columns), None)
    score_col = next((c for c in ["risk_score","reliability_score"] if c in sc.columns), None)
    rev_col   = next((c for c in ["total_revenue_at_risk","stockout_revenue_lost"] if c in sc.columns), None)
    name_col  = next((c for c in ["supplier_name"] if c in sc.columns), None)
    lead_col  = next((c for c in ["avg_lead_actual","lead_time_days_avg"] if c in sc.columns), None)

    # ── KPI strip ─────────────────────────────────────────────
    high_n = int((sc[risk_col].str.contains("HIGH",na=False)).sum()) if risk_col else 0
    med_n  = int((sc[risk_col].str.contains("MEDIUM",na=False)).sum()) if risk_col else 0
    low_n  = int((sc[risk_col].str.contains("LOW",na=False)).sum()) if risk_col else 0
    avg_f  = round(float(sc[fill_col].mean())*100,1) if fill_col else 0
    avg_l  = round(float(sc[lead_col].mean()),1) if lead_col else 0
    tot_st = int(sc[stock_col].sum()) if stock_col else 0
    kpi_row([("Total Suppliers",f"{len(sc)}","",C_TEAL),
             ("Avg Fill Rate",f"{avg_f:.1f}%","Network average",C_GREEN if avg_f>=85 else C_RED),
             ("HIGH RISK",f"{high_n}","Immediate action",C_RED),
             ("MEDIUM RISK",f"{med_n}","Monitor closely",C_AMBER),
             ("LOW RISK",f"{low_n}","Performing well",C_GREEN),
             ("Total Stockouts",f"{tot_st:,}","Caused by suppliers",C_PURPLE)])

    # ── Supplier selector ─────────────────────────────────────
    sup_opts = (["All suppliers"] + sorted(sc[name_col].dropna().tolist())) if name_col else ["All suppliers"]
    sel_sup = st.selectbox("Select supplier for deep dive:",sup_opts,key="sup_sel_main")
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Charts ────────────────────────────────────────────────
    if sel_sup == "All suppliers":
        sc1,sc2 = st.columns(2)
        with sc1:
            sh("Fill Rate vs Reliability — by Risk Tier")
            if fill_col and score_col and not sc.empty:
                tier_colors = {"HIGH_RISK":C_RED,"MEDIUM_RISK":C_AMBER,"LOW_RISK":C_GREEN}
                fig_s = go.Figure()
                for tier, color in tier_colors.items():
                    sub = sc[sc[risk_col]==tier] if risk_col else sc
                    if not sub.empty:
                        fig_s.add_trace(go.Scatter(
                            x=sub[score_col], y=sub[fill_col]*100,
                            mode="markers", name=tier.replace("_"," "),
                            marker=dict(color=color, size=10, opacity=0.8),
                            text=sub[name_col] if name_col else None,
                            hovertemplate="%{text}<br>Score: %{x:.1f}<br>Fill: %{y:.1f}%<extra></extra>"))
                fig_s.add_hline(y=85,line_dash="dot",line_color=C_AMBER,
                    annotation_text="85% fill rate target",
                    annotation_font=dict(color=C_AMBER,size=9))
                fig_s.update_layout(**DL("",height=340))
                fig_s.update_xaxes(title="Risk Score")
                fig_s.update_yaxes(title="Fill Rate %")
                PC(fig_s,"sup_scatter")
            else: PC(EF(height=340),"sup_scatter_e")
        with sc2:
            sh("Stockout Events by Supplier — Top 15")
            if stock_col and name_col and not sc.empty:
                top_s = sc.nlargest(15,stock_col)
                fig_st = go.Figure(go.Bar(
                    x=top_s[stock_col],y=top_s[name_col],orientation="h",
                    marker=dict(color=top_s[stock_col],
                        colorscale=[[0,C_AMBER],[1,C_RED]],showscale=False),
                    text=[f"{int(v):,}" for v in top_s[stock_col]],
                    textposition="outside",textfont=dict(color=C_WHITE,size=8)))
                fig_st.update_layout(**DL("",height=340))
                PC(fig_st,"sup_stock")
            else: PC(EF(height=340),"sup_stock_e")
    else:
        # Single supplier deep dive charts
        sup_row = sc[sc[name_col]==sel_sup].iloc[0] if name_col and not sc[sc[name_col]==sel_sup].empty else None
        if sup_row is not None and fill_col and late_col and short_col and stock_col:
            max_st = sc[stock_col].max()
            vals = [
                float(sup_row[fill_col])*100,
                100-float(sup_row[late_col])*100,
                100-float(sup_row[short_col])*100,
                100-float(sup_row[stock_col])/max(max_st,1)*100
            ]
            cats = ["Fill Rate","On-Time %","Full Delivery %","Low Stockout Score"]
            bar_colors = [C_GREEN if v>=80 else (C_AMBER if v>=60 else C_RED) for v in vals]
            fig_dv = go.Figure(go.Bar(x=cats,y=vals,marker_color=bar_colors,
                text=[f"{v:.1f}" for v in vals],textposition="outside",
                textfont=dict(color=C_WHITE,size=11)))
            fig_dv.add_hline(y=80,line_dash="dot",line_color=C_GREEN,
                annotation_text="Target 80+",annotation_font=dict(color=C_GREEN,size=9))
            fig_dv.update_layout(**DL(f"{sel_sup} — 4-Factor Scorecard",height=300))
            fig_dv.update_yaxes(range=[0,115])
            PC(fig_dv,"sup_dv_chart")
            tier = str(sup_row.get(risk_col,"")) if risk_col else ""
            tc = C_RED if "HIGH" in tier else (C_AMBER if "MEDIUM" in tier else C_GREEN)
            m1,m2,m3,m4 = st.columns(4)
            with m1: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_GREEN};border-radius:8px;padding:12px;text-align:center"><div style="font-size:20px;font-weight:800;color:{C_GREEN}">{float(sup_row[fill_col])*100:.1f}%</div><div style="font-size:10px;color:{C_GRAY}">Fill Rate</div></div>', unsafe_allow_html=True)
            with m2: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_RED};border-radius:8px;padding:12px;text-align:center"><div style="font-size:20px;font-weight:800;color:{C_RED}">{float(sup_row[late_col])*100:.1f}%</div><div style="font-size:10px;color:{C_GRAY}">Late Deliveries</div></div>', unsafe_allow_html=True)
            with m3: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_AMBER};border-radius:8px;padding:12px;text-align:center"><div style="font-size:20px;font-weight:800;color:{C_AMBER}">{float(sup_row[short_col])*100:.1f}%</div><div style="font-size:10px;color:{C_GRAY}">Short Deliveries</div></div>', unsafe_allow_html=True)
            with m4: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {tc};border-radius:8px;padding:12px;text-align:center"><div style="font-size:20px;font-weight:800;color:{tc}">{tier}</div><div style="font-size:10px;color:{C_GRAY}">Risk Tier</div></div>', unsafe_allow_html=True)

    # ── Scorecard table ───────────────────────────────────────
    st.markdown('<div class="card">', unsafe_allow_html=True)
    sh("Full Supplier Scorecard","Sorted by fill rate — click column headers to sort")
    show_c = [c for c in ["supplier_name","risk_tier","risk_score","avg_fulfillment_rate",
                          "late_delivery_rate","short_delivery_rate","avg_lead_actual",
                          "stockout_events_caused","total_revenue_at_risk","total_orders"] if c in sc.columns]
    if show_c:
        sc_display = sc[show_c].copy()
        if "avg_fulfillment_rate" in sc_display.columns:
            sc_display["avg_fulfillment_rate"] = (sc_display["avg_fulfillment_rate"]*100).round(1)
        if "late_delivery_rate" in sc_display.columns:
            sc_display["late_delivery_rate"] = (sc_display["late_delivery_rate"]*100).round(1)
        if "short_delivery_rate" in sc_display.columns:
            sc_display["short_delivery_rate"] = (sc_display["short_delivery_rate"]*100).round(1)
        sort_by = "avg_fulfillment_rate" if "avg_fulfillment_rate" in sc_display.columns else show_c[0]
        st.dataframe(sc_display.sort_values(sort_by).rename(columns={
            "supplier_name":"Supplier","risk_tier":"Risk","risk_score":"Score",
            "avg_fulfillment_rate":"Fill Rate %","late_delivery_rate":"Late %",
            "short_delivery_rate":"Short %","avg_lead_actual":"Lead Days",
            "stockout_events_caused":"Stockouts","total_revenue_at_risk":"Revenue at Risk",
            "total_orders":"Orders"}),
            use_container_width=True,height=340,hide_index=True)
    st.markdown("</div>", unsafe_allow_html=True)

    # ── Supplier AI ────────────────────────────────────────────
    st.markdown("<hr>", unsafe_allow_html=True)
    sh("AI Assistant — Supplier Intelligence","")
    try:
        import importlib, sys as _sys2
        _sp2 = str(Path(__file__).resolve().parent)
        if _sp2 not in _sys2.path: _sys2.path.insert(0, _sp2)
        import supplier_ai as _sai
        importlib.reload(_sai)
        _sai_ok = True
    except Exception as _se:
        _sai_ok = False
    _sup_key = "supplier"
    if _sup_key not in st.session_state["chat"]:
        st.session_state["chat"][_sup_key] = []
    _sup_msgs = st.session_state["chat"][_sup_key]
    _sup_quick = [
        ("Dual-source now","Which suppliers should I dual-source immediately?"),
        ("Most stockouts","Which supplier is causing the most stockouts?"),
        ("HIGH RISK detail","Show me all HIGH RISK suppliers with full breakdown"),
        ("Selected supplier",f"Deep dive into {sel_sup}" if sel_sup != "All suppliers" else "Select a supplier above"),
    ]
    _sqcols = st.columns(len(_sup_quick))
    for _sqi,(_sql,_sqq) in enumerate(_sup_quick):
        with _sqcols[_sqi]:
            if st.button(_sql,key=f"qq_sup_{_sqi}",use_container_width=True):
                st.session_state["_pend_sup"] = _sqq
    for _m in _sup_msgs:
        if _m["role"]=="user":
            _uh = '<div class="chat-label" style="text-align:right">You</div>'
            _uh += '<div class="chat-user">' + str(_m["content"]) + '</div>'
            st.markdown(_uh, unsafe_allow_html=True)
        else:
            _stxt = str(_m["content"]).replace("<","&lt;").replace(">","&gt;").replace(chr(10),"<br>")
            _ah = '<div class="chat-label">AI</div><div class="chat-ai">' + _stxt + '</div>'
            st.markdown(_ah, unsafe_allow_html=True)
            if _m.get("fig") is not None:
                try:
                    st.plotly_chart(_m["fig"],use_container_width=True,
                        config={"displayModeBar":False},
                        key="sup_chart_"+str(_sup_msgs.index(_m)))
                except: pass
    if not _sup_msgs:
        st.markdown('<div class="chat-ai">Select a supplier above for a deep dive, or ask: dual-source · most stockouts · HIGH RISK suppliers</div>', unsafe_allow_html=True)
    _sup_pend = st.session_state.pop("_pend_sup", None)
    _sup_in = st.chat_input("Ask about suppliers...", key="ci_supplier")
    _sup_in = _sup_in or _sup_pend
    if _sup_in:
        _sup_msgs.append({"role":"user","content":_sup_in,"fig":None})
        with st.spinner("Analysing supplier data..."):
            if _sai_ok:
                _selected = sel_sup if sel_sup != "All suppliers" else None
                _stxt2, _sfig2 = _sai.answer(_sup_in, sc, _selected)
            else:
                _stxt2 = "supplier_ai module not loaded."
                _sfig2 = None
        _sup_msgs.append({"role":"assistant","content":_stxt2,"fig":_sfig2})
        st.rerun()

# ══════════════════════════════════════════════════════════════
# FORECAST & MODEL — Bug 1 fixed: fillcolor uses rgba()
# ══════════════════════════════════════════════════════════════
elif PAGE == "Forecast & Model":
    sh("Forecast & Model","DemandSense v2 · weekly accuracy · demand ranges · stockout prediction")
    st.markdown("<hr>", unsafe_allow_html=True)
    ms=D["model_sum"]; wm=D["weekly"]; acc=D["acc_cat"]; pred=D["pred"]
    v2r=ms[ms["forecast_method"].str.contains("DemandSense_v2",na=False)].iloc[0] if not ms.empty else None
    b30=ms[ms["forecast_method"].str.contains("MovingAvg30",na=False)].iloc[0] if not ms.empty else None
    if v2r is not None:
        kpi_row([
            ("WAPE",f"{float(v2r['wape']):.2f}%",f"vs {float(b30['wape']) if b30 is not None else 28.89:.2f}% baseline",C_TEAL),
            ("P90 Coverage",f"{float(v2r.get('p90_coverage',0)):.2f}%","Target 88-92%",C_GREEN),
            ("MAE",f"{float(v2r.get('mae',0)):.2f}u","Mean abs error",C_AMBER),
            ("Bias",f"{float(v2r.get('bias',0)):.3f}","Near zero = perfect",C_PURPLE),
        ])

    SPIKE_REASONS = {
        "2025-11-24/2025-11-30": ("Black Friday week",
            "Model over-predicted demand. Stores cleared stock fast — actual sales hit zero while forecast stayed high. "
            "Bias -1.72 means for every 10 units sold, model predicted 17. Next week corrected sharply."),
        "2025-12-22/2025-12-28": ("Christmas week",
            "Christmas demand surge then sharp drop after Dec 25. "
            "Model saw pre-holiday velocity and kept forecasting high — actual sales collapsed post-holiday. Bias -1.25."),
        "2025-12-29/2026-01-04": ("Post-Christmas week",
            "Post-holiday demand collapse. Model over-predicted because of high December training signal. "
            "Stores went quiet while model expected continued volume. Worst bias of all weeks at -1.82."),
    }

    tab1,tab2,tab3,tab4,tab5 = st.tabs([
        "Weekly Monitoring","Category Accuracy","Demand Range P10-P90","Forecast vs Actual","Backtest"])

    with tab1:
        if not wm.empty:
            avg_wape = wm["wape"].mean()
            wm2 = wm.copy()
            wm2["is_spike"] = wm2["wape"] > avg_wape * 1.2
            tc1,tc2 = st.columns([3,1])
            with tc1:
                colors_ = [C_RED if r["is_spike"] else C_TEAL for _,r in wm2.iterrows()]
                fig_=go.Figure()
                fig_.add_trace(go.Bar(x=wm2["week"],y=wm2["wape"],name="WAPE %",
                    marker_color=colors_,
                    text=[f"{v:.1f}%" for v in wm2["wape"]],
                    textposition="outside",textfont=dict(color=C_WHITE,size=9)))
                if "bias" in wm2.columns:
                    fig_.add_trace(go.Scatter(x=wm2["week"],y=wm2["bias"],
                        mode="lines+markers",name="Bias",
                        line=dict(color=C_AMBER,width=2,dash="dot"),
                        yaxis="y2",marker=dict(size=6,color=C_AMBER)))
                fig_.add_hrect(y0=22,y1=26,fillcolor="rgba(16,185,129,0.10)",
                    opacity=1,line_width=0,
                    annotation_text="Target 22-26%",
                    annotation_font=dict(color=C_GREEN,size=9))
                fig_.update_layout(**DL("Weekly WAPE (red=spike) + Bias (dashed amber)",height=360,
                    xaxis=dict(tickangle=-45),
                    yaxis2=dict(title="Bias",overlaying="y",side="right",
                        showgrid=False,tickfont=dict(color=C_AMBER),range=[-3,3])))
                PC(fig_,"fc_weekly")
            with tc2:
                week_opts = wm2["week"].tolist()
                st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:6px">Select week to investigate</div>', unsafe_allow_html=True)
                sel_week = st.selectbox("Week",week_opts,index=len(week_opts)-1,
                    label_visibility="collapsed",key="fc_week_sel")
                sel_row = wm2[wm2["week"]==sel_week].iloc[0] if not wm2[wm2["week"]==sel_week].empty else None
                if sel_row is not None:
                    wv=float(sel_row["wape"]); bv=float(sel_row["bias"])
                    is_spike=bool(sel_row["is_spike"])
                    sc=C_RED if is_spike else C_GREEN
                    lbl="SPIKE" if is_spike else "NORMAL"
                    st.markdown(f'<div style="background:{C_CARD};border:1px solid {sc}44;border-radius:8px;padding:12px;margin-bottom:8px"><div style="font-size:9px;color:{sc};font-weight:700;text-transform:uppercase;margin-bottom:4px">{lbl}</div><div style="font-size:22px;font-weight:800;color:{sc}">{wv:.1f}%</div><div style="font-size:10px;color:{C_GRAY}">WAPE</div><hr style="border-color:{C_BORDER};margin:8px 0"><div style="font-size:14px;font-weight:700;color:{C_AMBER}">{bv:+.3f}</div><div style="font-size:10px;color:{C_GRAY}">Bias (- = over-predicted)</div></div>', unsafe_allow_html=True)
                    if sel_week in SPIKE_REASONS:
                        name_,reason_=SPIKE_REASONS[sel_week]
                        st.markdown(f'<div style="background:{C_CARD};border-left:3px solid {C_RED};border-radius:0 8px 8px 0;padding:12px"><div style="font-size:11px;font-weight:700;color:{C_RED};margin-bottom:6px">{name_}</div><div style="font-size:11px;color:{C_GRAY};line-height:1.6">{reason_}</div></div>', unsafe_allow_html=True)
                    elif is_spike:
                        st.markdown(f'<div style="background:{C_CARD};border-left:3px solid {C_AMBER};padding:10px 12px;border-radius:0 8px 8px 0;font-size:11px;color:{C_GRAY}">Above-average WAPE. Bias {bv:+.2f} — {"model over-predicted" if bv<0 else "model under-predicted"}.</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div style="background:{C_CARD};border-left:3px solid {C_GREEN};padding:10px 12px;border-radius:0 8px 8px 0;font-size:11px;color:{C_GRAY}">Normal week. WAPE {wv:.1f}% within target band. Bias {bv:+.3f} — model well calibrated.</div>', unsafe_allow_html=True)
            st.markdown('<div class="card">', unsafe_allow_html=True)
            wmd=wm2.copy()
            wmd["status"]=wmd["is_spike"].map({True:"SPIKE",False:"Normal"})
            wmd["wape"]=wmd["wape"].round(2); wmd["bias"]=wmd["bias"].round(3)
            wmd["p90_coverage"]=wmd["p90_coverage"].round(2)
            st.dataframe(wmd[["week","wape","bias","p90_coverage","status"]].rename(
                columns={"wape":"WAPE %","bias":"Bias","p90_coverage":"P90 Cov %","status":"Status"}),
                use_container_width=True,height=280,hide_index=True)
            st.markdown("</div>", unsafe_allow_html=True)
        else: PC(EF(height=340),"fc_weekly_e")

    with tab2:
        if not acc.empty and "category" in acc.columns and "wape" in acc.columns:
            t2c1,t2c2=st.columns([2,1])
            with t2c1:
                acc_s=acc.sort_values("wape",ascending=True)
                fig2_=go.Figure(go.Bar(x=acc_s["wape"],y=acc_s["category"],orientation="h",
                    marker=dict(color=acc_s["wape"],
                        colorscale=[[0,C_GREEN],[0.5,C_AMBER],[1,C_RED]],showscale=True,
                        colorbar=dict(title=dict(text="WAPE %",font=dict(color=C_WHITE)),
                            tickfont=dict(color=C_WHITE))),
                    text=[f"{v:.1f}%" for v in acc_s["wape"]],textposition="outside",
                    textfont=dict(color=C_WHITE,size=9)))
                fig2_.update_layout(**DL("WAPE by Product Category",height=400))
                PC(fig2_,"fc_cat")
            with t2c2:
                best_=acc.loc[acc["wape"].idxmin()]
                worst_=acc.loc[acc["wape"].idxmax()]
                st.markdown(f'<div style="background:{C_CARD};border:1px solid {C_GREEN}44;border-radius:8px;padding:14px;margin-bottom:10px"><div style="font-size:9px;color:{C_GREEN};font-weight:700;text-transform:uppercase;margin-bottom:4px">Best forecast</div><div style="font-size:14px;font-weight:700;color:{C_WHITE}">{best_["category"]}</div><div style="font-size:22px;font-weight:800;color:{C_GREEN}">{best_["wape"]:.1f}%</div><div style="font-size:10px;color:{C_GRAY}">{int(best_.get("rows",0)):,} training rows</div></div>', unsafe_allow_html=True)
                st.markdown(f'<div style="background:{C_CARD};border:1px solid {C_RED}44;border-radius:8px;padding:14px"><div style="font-size:9px;color:{C_RED};font-weight:700;text-transform:uppercase;margin-bottom:4px">Needs improvement</div><div style="font-size:14px;font-weight:700;color:{C_WHITE}">{worst_["category"]}</div><div style="font-size:22px;font-weight:800;color:{C_RED}">{worst_["wape"]:.1f}%</div><div style="font-size:10px;color:{C_GRAY}">{int(worst_.get("rows",0)):,} training rows</div></div>', unsafe_allow_html=True)
        else: PC(EF(height=380),"fc_cat_e")

    with tab3:
        sh("Demand Range","P10 = low scenario · P50 = median · P90 = order at this level to avoid stockout")
        SKUS_FC=sorted(pred["sku_id"].dropna().unique().tolist()) if not pred.empty else []
        r3c1,r3c2=st.columns([2,2])
        with r3c1:
            st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store</div>', unsafe_allow_html=True)
            fc_store_r=st.selectbox("Store r",[S_NAMES.get(s,s) for s in STORES[:80]],label_visibility="collapsed",key="fc_range_store")
        with r3c2:
            st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">SKU</div>', unsafe_allow_html=True)
            fc_sku_r=st.selectbox("SKU r",SKUS_FC[:100],label_visibility="collapsed",key="fc_range_sku")
        sid_r=fc_store_r.split(" — ")[0] if fc_store_r else None
        if not pred.empty and sid_r and fc_sku_r:
            dr=pred[(pred["store_id"]==sid_r)&(pred["sku_id"]==fc_sku_r)].copy().sort_values("date")
            if not dr.empty and "lower_bound_90" in dr.columns:
                p50_=dr["forecast_units"].mean()
                p10_=dr["lower_bound_90"].mean()
                p90_=dr["upper_bound_90"].mean()
                band_=p90_-p10_
                unc_="HIGH" if band_>p50_*0.5 else ("MEDIUM" if band_>p50_*0.25 else "LOW")
                uc_=C_RED if unc_=="HIGH" else (C_AMBER if unc_=="MEDIUM" else C_GREEN)
                fig_r=go.Figure()
                fig_r.add_trace(go.Scatter(
                    x=list(dr["date"])+list(dr["date"])[::-1],
                    y=list(dr["upper_bound_90"])+list(dr["lower_bound_90"])[::-1],
                    fill="toself",fillcolor="rgba(20,184,166,0.12)",
                    line=dict(color="rgba(0,0,0,0)"),name="P10-P90 range"))
                fig_r.add_trace(go.Scatter(x=dr["date"],y=dr["forecast_units"],name="P50 forecast",
                    line=dict(color=C_TEAL,width=2.5)))
                fig_r.add_trace(go.Scatter(x=dr["date"],y=dr["upper_bound_90"],name="P90 — order target",
                    line=dict(color=C_AMBER,width=1.5,dash="dash")))
                fig_r.add_trace(go.Scatter(x=dr["date"],y=dr["lower_bound_90"],name="P10 — low scenario",
                    line=dict(color=C_GRAY,width=1,dash="dot")))
                if "target" in dr.columns:
                    fig_r.add_trace(go.Scatter(x=dr["date"],y=dr["target"],name="Actual",
                        mode="markers",marker=dict(color=C_RED,size=6)))
                fig_r.update_layout(**DL(f"Demand Range: {fc_sku_r} at {sid_r}",height=340))
                fig_r.update_xaxes(title="Date"); fig_r.update_yaxes(title="Units/day")
                PC(fig_r,"fc_range_chart")
                rc1,rc2,rc3,rc4=st.columns(4)
                with rc1: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_GRAY};border-radius:8px;padding:12px;text-align:center"><div style="font-size:18px;font-weight:800;color:{C_GRAY}">{p10_:.1f}u</div><div style="font-size:10px;color:{C_GRAY}">P10 low scenario</div></div>', unsafe_allow_html=True)
                with rc2: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_TEAL};border-radius:8px;padding:12px;text-align:center"><div style="font-size:18px;font-weight:800;color:{C_TEAL}">{p50_:.1f}u</div><div style="font-size:10px;color:{C_GRAY}">P50 median expected</div></div>', unsafe_allow_html=True)
                with rc3: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_AMBER};border-radius:8px;padding:12px;text-align:center"><div style="font-size:18px;font-weight:800;color:{C_AMBER}">{p90_:.1f}u</div><div style="font-size:10px;color:{C_GRAY}">P90 — order at this</div></div>', unsafe_allow_html=True)
                with rc4: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {uc_};border-radius:8px;padding:12px;text-align:center"><div style="font-size:18px;font-weight:800;color:{uc_}">{unc_}</div><div style="font-size:10px;color:{C_GRAY}">Uncertainty · {band_:.1f}u band</div></div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div style="color:{C_GRAY};padding:20px">No range data for {fc_sku_r} at {sid_r}.</div>', unsafe_allow_html=True)
        st.markdown("<hr>", unsafe_allow_html=True)
        sh("Highest uncertainty SKUs","Widest P10-P90 band — most risky to under-order")
        if not pred.empty and "lower_bound_90" in pred.columns:
            ut=pred.copy(); ut["band"]=ut["upper_bound_90"]-ut["lower_bound_90"]
            ut_top=(ut.groupby(["store_id","sku_id","category"])
                .agg(p50=("forecast_units","mean"),p90=("upper_bound_90","mean"),
                     p10=("lower_bound_90","mean"),band=("band","mean"))
                .reset_index().nlargest(20,"band"))
            ut_top["unc_pct"]=(ut_top["band"]/ut_top["p50"]*100).round(1)
            for c in ["p50","p90","p10","band"]: ut_top[c]=ut_top[c].round(1)
            st.dataframe(ut_top[["store_id","sku_id","category","p10","p50","p90","band","unc_pct"]].rename(
                columns={"store_id":"Store","sku_id":"SKU","category":"Category",
                    "p10":"P10","p50":"P50","p90":"P90","band":"Band Width","unc_pct":"Uncertainty %"}),
                use_container_width=True,height=300,hide_index=True)

    with tab4:
        sh("Forecast vs Actual","What the model predicted vs what actually sold")
        SKUS_FC2=sorted(pred["sku_id"].dropna().unique().tolist()) if not pred.empty else []
        f4c1,f4c2=st.columns(2)
        with f4c1:
            st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store</div>', unsafe_allow_html=True)
            fva_store=st.selectbox("Store fva",[S_NAMES.get(s,s) for s in STORES[:80]],label_visibility="collapsed",key="fc_fva_store")
        with f4c2:
            st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">SKU</div>', unsafe_allow_html=True)
            fva_sku=st.selectbox("SKU fva",SKUS_FC2[:100],label_visibility="collapsed",key="fc_fva_sku")
        fva_sid=fva_store.split(" — ")[0] if fva_store else None
        if not pred.empty and fva_sid and fva_sku and "target" in pred.columns:
            fva=pred[(pred["store_id"]==fva_sid)&(pred["sku_id"]==fva_sku)].copy().sort_values("date")
            if not fva.empty:
                mae_=(fva["forecast_units"]-fva["target"]).abs().mean()
                bias_=(fva["forecast_units"]-fva["target"]).mean()
                wape_=(fva["forecast_units"]-fva["target"]).abs().sum()/max(fva["target"].sum(),1)*100
                fig_fva=go.Figure()
                fig_fva.add_trace(go.Scatter(x=fva["date"],y=fva["target"],name="Actual",
                    line=dict(color=C_RED,width=2),mode="lines+markers",marker=dict(size=5)))
                fig_fva.add_trace(go.Scatter(x=fva["date"],y=fva["forecast_units"],name="Forecast P50",
                    line=dict(color=C_TEAL,width=2.5,dash="dash")))
                if "upper_bound_90" in fva.columns:
                    fig_fva.add_trace(go.Scatter(
                        x=list(fva["date"])+list(fva["date"])[::-1],
                        y=list(fva["upper_bound_90"])+list(fva["lower_bound_90"])[::-1],
                        fill="toself",fillcolor="rgba(20,184,166,0.10)",
                        line=dict(color="rgba(0,0,0,0)"),name="P10-P90 band"))
                fig_fva.update_layout(**DL(f"Forecast vs Actual: {fva_sku} at {fva_sid}",height=340))
                fig_fva.update_xaxes(title="Date"); fig_fva.update_yaxes(title="Units/day")
                PC(fig_fva,"fc_fva_chart")
                fvc1,fvc2,fvc3=st.columns(3)
                with fvc1: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_AMBER};border-radius:8px;padding:12px;text-align:center"><div style="font-size:20px;font-weight:800;color:{C_AMBER}">{wape_:.1f}%</div><div style="font-size:10px;color:{C_GRAY}">WAPE for this SKU</div></div>', unsafe_allow_html=True)
                with fvc2:
                    bc=C_RED if abs(bias_)>2 else C_GREEN
                    st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {bc};border-radius:8px;padding:12px;text-align:center"><div style="font-size:20px;font-weight:800;color:{bc}">{bias_:+.1f}u</div><div style="font-size:10px;color:{C_GRAY}">Bias (- = over-predicted)</div></div>', unsafe_allow_html=True)
                with fvc3: st.markdown(f'<div style="background:{C_CARD};border-top:3px solid {C_TEAL};border-radius:8px;padding:12px;text-align:center"><div style="font-size:20px;font-weight:800;color:{C_TEAL}">{mae_:.1f}u</div><div style="font-size:10px;color:{C_GRAY}">MAE avg daily error</div></div>', unsafe_allow_html=True)

    with tab5:
        bt=D["backtest_s"]
        if not bt.empty and "policy" in bt.columns:
            old_=bt[bt["policy"]=="OLD"].iloc[0] if not bt[bt["policy"]=="OLD"].empty else None
            new_=bt[bt["policy"]=="NEW"].iloc[0] if not bt[bt["policy"]=="NEW"].empty else None
            if old_ is not None and new_ is not None:
                so_old=max(old_["stockout_events"],0); so_new=max(new_["stockout_events"],0)
                so_red=(so_old-so_new)/max(so_old,1)*100
                lu_red=(float(old_["lost_units"])-float(new_["lost_units"]))/max(float(old_["lost_units"]),1)*100
                sl_old=float(old_.get("service_level_pct",0)); sl_new=float(new_.get("service_level_pct",0))
                kpi_row([("Stockout Reduction",f"{so_red:.1f}%","OLD vs AI policy",C_GREEN),
                         ("Lost Units Reduction",f"{lu_red:.1f}%","OLD vs AI policy",C_GREEN),
                         ("OLD Stockouts",f"{int(so_old):,}","Moving avg baseline",C_RED),
                         ("NEW Stockouts",f"{int(so_new):,}","AI DemandSense v2",C_GREEN),
                         ("OLD Service Level",f"{sl_old:.1f}%","Baseline",C_RED),
                         ("NEW Service Level",f"{sl_new:.1f}%","AI policy",C_GREEN)])
                if so_old==0 and so_new==0:
                    st.markdown(f'<div style="background:{C_CARD};border:1px solid {C_BORDER};border-radius:8px;padding:12px;font-size:12px;color:{C_GRAY};margin-top:8px"><strong style="color:{C_TEAL2}">Note:</strong> Both policies show 0 stockouts in this {int(old_["demand_units"]):,}-unit sample. The full Black Friday simulation showed 30%+ reduction — this sample covers a well-stocked period.</div>', unsafe_allow_html=True)
        else: PC(EF("Run python src/pipeline.py to generate backtest data",300),"fc_bt_e")

    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="card-glow-r">', unsafe_allow_html=True)
    sh("Stores stocking out before this weekend","Products with 2 days or less of supply")
    na_a=D["na"]
    if not na_a.empty and "days_of_supply_current" in na_a.columns:
        wr=na_a[na_a["days_of_supply_current"]<=2].copy()
        if not wr.empty:
            wr_rev=float(wr["revenue_at_risk"].sum()) if "revenue_at_risk" in wr.columns else 0
            st.markdown(f'<div style="font-size:13px;color:{C_WHITE};margin-bottom:12px"><strong style="color:{C_RED}">{len(wr):,} products</strong> across <strong style="color:{C_RED}">{wr["store_id"].nunique()} stores</strong> — <strong style="color:{C_RED}">${wr_rev:,.0f}</strong> at risk this weekend</div>', unsafe_allow_html=True)
            sc_wr=[c for c in ["store_id","sku_id","product_name","category","days_of_supply_current","units_to_order","revenue_at_risk","supplier_name"] if c in wr.columns]
            st.dataframe(wr.sort_values("days_of_supply_current")[sc_wr].head(50),use_container_width=True,height=280,hide_index=True)
        else:
            st.markdown(f'<div style="color:{C_GREEN};padding:16px;font-size:14px;font-weight:700">No products stocking out this weekend</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div style="color:{C_GRAY};padding:16px">Alert data not loaded.</div>', unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)
    sh("Store vs Store Forecast Comparison","Select two stores and a SKU to compare demand forecasts side by side")
    SKUS_CMP=sorted(pred["sku_id"].dropna().unique().tolist()) if not pred.empty else []
    cfc1,cfc2,cfc3=st.columns(3)
    with cfc1:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store A</div>', unsafe_allow_html=True)
        cmp_store_a=st.selectbox("Store A",[S_NAMES.get(s,s) for s in STORES[:80]],label_visibility="collapsed",key="fc_store_a")
    with cfc2:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Store B</div>', unsafe_allow_html=True)
        cmp_store_b=st.selectbox("Store B",[S_NAMES.get(s,s) for s in STORES[:80]],index=min(1,len(STORES)-1),label_visibility="collapsed",key="fc_store_b")
    with cfc3:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Product SKU</div>', unsafe_allow_html=True)
        cmp_sku=st.selectbox("SKU cmp",SKUS_CMP[:100],label_visibility="collapsed",key="fc_sku")
    sid_a=cmp_store_a.split(" — ")[0] if cmp_store_a else None
    sid_b=cmp_store_b.split(" — ")[0] if cmp_store_b else None
    if not pred.empty and cmp_sku and sid_a and sid_b:
        da=pred[(pred["store_id"]==sid_a)&(pred["sku_id"]==cmp_sku)].copy().sort_values("date")
        db=pred[(pred["store_id"]==sid_b)&(pred["sku_id"]==cmp_sku)].copy().sort_values("date")
        if not da.empty or not db.empty:
            fig_c=go.Figure()
            if not da.empty:
                fig_c.add_trace(go.Scatter(x=da["date"],y=da["forecast_units"],name=sid_a,line=dict(color=C_TEAL,width=2.5)))
                if "upper_bound_90" in da.columns:
                    fig_c.add_trace(go.Scatter(x=list(da["date"])+list(da["date"])[::-1],
                        y=list(da["upper_bound_90"])+list(da["lower_bound_90"])[::-1],
                        fill="toself",fillcolor="rgba(13,148,136,0.10)",
                        line=dict(color="rgba(0,0,0,0)"),name=f"{sid_a} P10-P90"))
            if not db.empty:
                fig_c.add_trace(go.Scatter(x=db["date"],y=db["forecast_units"],name=sid_b,line=dict(color=C_AMBER,width=2.5,dash="dash")))
                if "upper_bound_90" in db.columns:
                    fig_c.add_trace(go.Scatter(x=list(db["date"])+list(db["date"])[::-1],
                        y=list(db["upper_bound_90"])+list(db["lower_bound_90"])[::-1],
                        fill="toself",fillcolor="rgba(245,158,11,0.10)",
                        line=dict(color="rgba(0,0,0,0)"),name=f"{sid_b} P10-P90"))
            fig_c.update_layout(**DL(f"Forecast Comparison: {cmp_sku}",height=320))
            fig_c.update_xaxes(title="Date"); fig_c.update_yaxes(title="Units/day")
            PC(fig_c,"fc_cmp_chart")
            ca_,cb_=st.columns(2)
            with ca_:
                if not da.empty:
                    avg_a=da["forecast_units"].mean()
                    p90_a=da["upper_bound_90"].mean() if "upper_bound_90" in da.columns else avg_a
                    st.markdown(f'<div style="background:{C_CARD};border:1px solid {C_TEAL}44;border-radius:8px;padding:14px"><div style="color:{C_TEAL};font-size:11px;font-weight:700;margin-bottom:6px">{sid_a}</div><div style="font-size:22px;font-weight:800;color:{C_TEAL}">{avg_a:.1f}u/day</div><div style="font-size:11px;color:{C_GRAY}">P50 avg · Order at P90: {p90_a:.1f}u/day</div></div>', unsafe_allow_html=True)
            with cb_:
                if not db.empty:
                    avg_b=db["forecast_units"].mean()
                    p90_b=db["upper_bound_90"].mean() if "upper_bound_90" in db.columns else avg_b
                    st.markdown(f'<div style="background:{C_CARD};border:1px solid {C_AMBER}44;border-radius:8px;padding:14px"><div style="color:{C_AMBER};font-size:11px;font-weight:700;margin-bottom:6px">{sid_b}</div><div style="font-size:22px;font-weight:800;color:{C_AMBER}">{avg_b:.1f}u/day</div><div style="font-size:11px;color:{C_GRAY}">P50 avg · Order at P90: {p90_b:.1f}u/day</div></div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div style="color:{C_GRAY};padding:20px">No forecast data for {cmp_sku} at selected stores.</div>', unsafe_allow_html=True)

    # ── Forecast AI — direct Ollama call with real data ─────────
    # ── Forecast AI — True Hybrid via forecast_ai module ────────
    # Forecast AI
    st.markdown('<hr>', unsafe_allow_html=True)
    sh('AI Assistant — Forecast Intelligence', 'Ask anything about WAPE, accuracy, stockouts or demand uncertainty')
    try:
        import importlib, sys as _sys
        _sp = str(Path(__file__).resolve().parent)
        if _sp not in _sys.path: _sys.path.insert(0, _sp)
        import forecast_ai as _fai
        importlib.reload(_fai)
        _fai_ok = True
    except Exception as _fe:
        _fai_ok = False
    _fc_key = 'forecast'
    if _fc_key not in st.session_state['chat']:
        st.session_state['chat'][_fc_key] = []
    _fc_msgs = st.session_state['chat'][_fc_key]
    _fc_quick = [
        ('WAPE spikes', 'Why did WAPE spike during Black Friday and Post-Christmas?'),
        ('Worst category', 'Which category has the worst forecast accuracy and why?'),
        ('Weekend stockout', 'Which stores will stock out before this weekend?'),
        ('Demand uncertainty', 'Which SKUs have the widest demand uncertainty band?'),
    ]
    _qcols = st.columns(len(_fc_quick))
    for _qi, (_ql, _qq) in enumerate(_fc_quick):
        with _qcols[_qi]:
            if st.button(_ql, key=f'qq_fc_{_qi}', use_container_width=True):
                st.session_state['_pend_fc'] = _qq
    for _m in _fc_msgs:
        if _m['role'] == 'user':
            _uh = '<div class="chat-label" style="text-align:right">You</div>'
            _uh += '<div class="chat-user">' + str(_m['content']) + '</div>'
            st.markdown(_uh, unsafe_allow_html=True)
        else:
            _txt = str(_m['content']).replace('<','&lt;').replace('>','&gt;').replace(chr(10),'<br>')
            _ah = '<div class="chat-label">AI</div><div class="chat-ai">' + _txt + '</div>'
            st.markdown(_ah, unsafe_allow_html=True)
            if _m.get('fig') is not None:
                try:
                    st.plotly_chart(_m['fig'], use_container_width=True,
                        config={'displayModeBar':False},
                        key='fc_chart_' + str(_fc_msgs.index(_m)))
                except: pass
    if not _fc_msgs:
        st.markdown('<div class="chat-ai">Ask about WAPE spikes, category accuracy, demand uncertainty or weekend stockout risk.</div>', unsafe_allow_html=True)
    _fc_pend = st.session_state.pop('_pend_fc', None)
    _fc_in = st.chat_input('Ask about WAPE, accuracy, forecasts...', key='ci_forecast')
    _fc_in = _fc_in or _fc_pend
    if _fc_in:
        _fc_msgs.append({'role':'user','content':_fc_in,'fig':None})
        with st.spinner('Analysing your data...'):
            if _fai_ok:
                _txt2, _fig2 = _fai.answer(_fc_in, wm, acc, pred, D['na'], v2r, b30)
            else:
                _txt2 = 'forecast_ai module not loaded.'
                _fig2 = None
        _fc_msgs.append({'role':'assistant','content':_txt2,'fig':_fig2})
        st.rerun()
elif PAGE == "Export Center":
    sh("Export Center","Download datasets · region-filtered exports · Excel supplier scorecard")
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Region filter ─────────────────────────────────────────
    ef1,ef2 = st.columns([2,2])
    with ef1:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Filter by Region</div>', unsafe_allow_html=True)
        exp_regions = ["All regions"] + sorted(ns["region"].dropna().unique().tolist()) if not ns.empty and "region" in ns.columns else ["All regions"]
        exp_region = st.selectbox("exp_reg",exp_regions,label_visibility="collapsed",key="exp_region")
    with ef2:
        st.markdown(f'<div style="font-size:10px;color:{C_GRAY};text-transform:uppercase;font-weight:600;margin-bottom:4px">Filter by Alert Tier</div>', unsafe_allow_html=True)
        exp_tier = st.selectbox("exp_tier",["All tiers","CRITICAL","WARNING","MONITOR","OK"],label_visibility="collapsed",key="exp_tier")
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Buyer report shortcut ─────────────────────────────────
    sh("Buyer Report Shortcut","Pre-filtered CRITICAL alerts — ready to send")
    if not na.empty:
        buyer_df = na.copy()
        if exp_region != "All regions" and "region" in buyer_df.columns:
            buyer_df = buyer_df[buyer_df["region"]==exp_region]
        if exp_tier != "All tiers" and "alert_tier" in buyer_df.columns:
            buyer_df = buyer_df[buyer_df["alert_tier"]==exp_tier]
        elif "alert_tier" in buyer_df.columns:
            buyer_df = buyer_df[buyer_df["alert_tier"]=="CRITICAL"]
        buyer_df = buyer_df.sort_values("revenue_at_risk",ascending=False) if "revenue_at_risk" in buyer_df.columns else buyer_df
        rar_b = float(buyer_df["revenue_at_risk"].sum()) if "revenue_at_risk" in buyer_df.columns else 0
        region_tag = exp_region.replace(" ","_").lower() if exp_region != "All regions" else "all"
        tier_tag = exp_tier.lower() if exp_tier != "All tiers" else "critical"
        st.markdown(f'<div style="background:{C_CARD};border-left:4px solid {C_TEAL};border-radius:6px;padding:12px 16px;margin-bottom:12px"><div style="font-size:13px;font-weight:700;color:{C_WHITE}">{len(buyer_df):,} rows · {exp_region} · {exp_tier if exp_tier!="All tiers" else "CRITICAL"} alerts · ${rar_b/1e6:.1f}M at risk</div><div style="font-size:10px;color:{C_GRAY};margin-top:4px">store\_id, sku\_id, product\_name, category, alert\_tier, days\_of\_supply\_current, units\_to\_order, revenue\_at\_risk, supplier\_name, priority\_score</div></div>', unsafe_allow_html=True)
        st.download_button(
            f"Download {region_tag}_{tier_tag}_buyer_report.csv",
            data=buyer_df.to_csv(index=False),
            file_name=f"{region_tag}_{tier_tag}_buyer_report.csv",
            mime="text/csv", key="exp_buyer")
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Supplier scorecard Excel ──────────────────────────────
    sh("Supplier Scorecard Excel","Color-coded by risk tier · sortable columns")
    try:
        import openpyxl as _opxl
        from openpyxl import Workbook as _WB
        from openpyxl.styles import PatternFill as _PF, Font as _FT, Alignment as _AL
        from openpyxl.utils import get_column_letter as _gcl
        import io as _io
        if not D["sup_sc"].empty:
            _wb = _WB(); _ws = _wb.active; _ws.title="Supplier Scorecard"
            _hfill = _PF("solid",fgColor="0D9488")
            _fills = {"HIGH_RISK":_PF("solid",fgColor="EF4444"),"MEDIUM_RISK":_PF("solid",fgColor="F59E0B"),"LOW_RISK":_PF("solid",fgColor="10B981")}
            _sc_exp=D["sup_sc"]; _show = [c for c in ["supplier_name","risk_tier","risk_score","avg_fulfillment_rate","late_delivery_rate","short_delivery_rate","avg_lead_actual","stockout_events_caused","total_revenue_at_risk","stores_served"] if c in _sc_exp.columns]
            _lbls = {"supplier_name":"Supplier","risk_tier":"Risk Tier","risk_score":"Score","avg_fulfillment_rate":"Fill Rate","late_delivery_rate":"Late %","short_delivery_rate":"Short %","avg_lead_actual":"Lead Days","stockout_events_caused":"Stockouts","total_revenue_at_risk":"Rev at Risk","stores_served":"Stores"}
            for ci,col in enumerate(_show,1):
                c=_ws.cell(row=1,column=ci,value=_lbls.get(col,col))
                c.fill=_hfill; c.font=_FT(bold=True,color="FFFFFF",size=10)
                c.alignment=_AL(horizontal="center")
            _sc_s = _sc_exp.sort_values("risk_score",ascending=False) if "risk_score" in _sc_exp.columns else _sc_exp
            for ri,(_,row) in enumerate(_sc_s[_show].iterrows(),2):
                _t=str(row.get("risk_tier","")) if "risk_tier" in _show else ""
                _rf=_fills.get(_t,None)
                for ci,col in enumerate(_show,1):
                    val=row[col]
                    if col in ["avg_fulfillment_rate","late_delivery_rate","short_delivery_rate"]:
                        val=f"{float(val)*100:.1f}%" if str(val)!="nan" else ""
                    elif col=="total_revenue_at_risk": val=f"${float(val):,.0f}" if str(val)!="nan" else ""
                    c=_ws.cell(row=ri,column=ci,value=val)
                    if _rf: c.fill=_rf; c.font=_FT(color="FFFFFF",size=9)
            for ci,_ in enumerate(_show,1): _ws.column_dimensions[_gcl(ci)].width=16
            _ws.freeze_panes="A2"
            _buf=_io.BytesIO(); _wb.save(_buf)
            high_n=int((_sc_exp["risk_tier"]=="HIGH_RISK").sum()) if "risk_tier" in D["sup_sc"].columns else 0
            st.markdown(f'<div style="background:{C_CARD};border-left:4px solid {C_GREEN};border-radius:6px;padding:12px 16px;margin-bottom:12px"><div style="font-size:13px;font-weight:700;color:{C_WHITE}">{len(_sc_exp)} suppliers · {high_n} HIGH RISK · Color coded by tier</div><div style="font-size:10px;color:{C_GRAY};margin-top:4px">Red=HIGH RISK · Amber=MEDIUM · Green=LOW · Sorted by risk score</div></div>', unsafe_allow_html=True)
            st.download_button("Download supplier_scorecard.xlsx",data=_buf.getvalue(),
                file_name="supplier_scorecard.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="exp_sup_xl")
    except ImportError:
        st.markdown(f'<div style="color:{C_AMBER};padding:10px">openpyxl not installed. Run: pip install openpyxl</div>', unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Standard CSV exports ──────────────────────────────────
    sh("Downloads by Role","Pick the right file for your job")
    _ROLE_EXPORTS = {
        "Store Manager": [
            ("Morning Alert List — order these now",   NEXUS/"allstore"/"network_master_alerts.csv"),
            ("Aisle Check List — phantom inventory",   NEXUS/"allstore"/"phantom_confidence.csv"),
            ("Store KPI Summary",                      NEXUS/"allstore"/"network_store_summary.csv"),
        ],
        "Buyer / Category Manager": [
            ("Category Revenue Mismatch by Store",    NEXUS/"localization"/"store_category_mismatch.csv"),
            ("Store Localization Scores",              NEXUS/"localization"/"store_localization_scores.csv"),
            ("Category Network Mismatch",              NEXUS/"localization"/"category_network_mismatch.csv"),
            ("Supplier Scorecard CSV",                 NEXUS/"supplier"/"supplier_scorecard.csv"),
        ],
        "Supply Chain Planner": [
            ("Replenishment Policy — SS and ROP",     PROCESSED/"replenishment_policy_inputs_demandsense.csv"),
            ("Weekly WAPE Monitoring",                 PROCESSED/"weekly_monitor_demandsense.csv"),
            ("Phantom by Store Summary",               NEXUS/"allstore"/"phantom_by_store.csv"),
        ],
        "Executive": [
            ("ROI Executive Summary",                  NEXUS/"roi"/"roi_executive_summary.csv"),
        ],
        "Data Science / Model Team": [
            ("DemandSense v2 Predictions (404MB)",    PROCESSED/"demandSense_v2_predictions.csv"),
            ("Backtest Daily Results",                 PROCESSED/"backtest_daily_demandsense.csv"),
            ("Forecast Accuracy by Category",          NEXUS/"forecast"/"accuracy_by_category.csv"),
            ("Backtest Summary",                       PROCESSED/"backtest_summary_demandsense.csv"),
        ],
    }
    _role_colors = {"Store Manager":C_RED,"Buyer / Category Manager":C_AMBER,"Supply Chain Planner":C_TEAL,"Executive":C_GREEN,"Data Science / Model Team":C_GRAY}
    import os as _os_exp
    for _role, _exports in _ROLE_EXPORTS.items():
        _rc = _role_colors.get(_role, C_GRAY)
        st.markdown(f'<div style="font-size:11px;font-weight:700;color:{_rc};text-transform:uppercase;margin:16px 0 6px 0;letter-spacing:0.5px">{_role}</div>', unsafe_allow_html=True)
        for _label, _fpath in _exports:
            try:
                if _fpath.exists():
                    _size_mb = _os_exp.path.getsize(_fpath)/1e6
                    with open(_fpath,"rb") as _f2: _data2 = _f2.read()
                    _row_count = len(_data2.decode("utf-8","ignore").split(chr(10)))-2
                    _ec1,_ec2=st.columns([3,1])
                    with _ec1:
                        st.markdown(f'<div style="padding:5px 0;border-bottom:1px solid {C_BORDER}20"><span style="color:{C_WHITE};font-size:11px">{_label}</span><span style="color:{C_GRAY};font-size:10px;margin-left:12px">{_row_count:,} rows · {_size_mb:.1f} MB</span></div>', unsafe_allow_html=True)
                    with _ec2:
                        st.download_button("Download",data=_data2,file_name=_fpath.name,mime="text/csv",key=f"exp_{_fpath.stem}")
            except: pass
    st.markdown("<hr>", unsafe_allow_html=True)
    sh("AI Assistant — Export Intelligence")
    try:
        import importlib, sys as _sys10
        _sp10=str(Path(__file__).resolve().parent)
        if _sp10 not in _sys10.path: _sys10.path.insert(0,_sp10)
        import export_ai as _exai
        importlib.reload(_exai)
        _exai_ok=True
    except:
        _exai_ok=False
    _ex_key="export"
    if _ex_key not in st.session_state["chat"]: st.session_state["chat"][_ex_key]=[]
    _ex_msgs=st.session_state["chat"][_ex_key]
    _ex_quick=[
        ("Buyer report","Which file should I download for my buyer report?"),
        (f"{exp_region} CRITICAL",f"Download CRITICAL alerts for the {exp_region} region"),
        ("Supplier Excel","Export the supplier scorecard as Excel with risk colors"),
    ]
    _excols=st.columns(len(_ex_quick))
    for _exi,(_exl,_exq) in enumerate(_ex_quick):
        with _excols[_exi]:
            if st.button(_exl,key=f"qq_ex_{_exi}",use_container_width=True):
                st.session_state["_pend_ex"]=_exq
    for _m in _ex_msgs:
        if _m["role"]=="user":
            _uh='<div class="chat-label" style="text-align:right">You</div>'
            _uh+='<div class="chat-user">'+str(_m["content"])+'</div>'
            st.markdown(_uh,unsafe_allow_html=True)
        else:
            _extxt=str(_m["content"]).replace("<","&lt;").replace(">","&gt;").replace(chr(10),"<br>")
            _ah='<div class="chat-label">AI</div><div class="chat-ai">'+_extxt+'</div>'
            st.markdown(_ah,unsafe_allow_html=True)
            if _m.get("csv_bytes") is not None:
                st.download_button("Download "+str(_m.get("fname","file.csv")),
                    data=_m["csv_bytes"],file_name=str(_m.get("fname","file.csv")),
                    mime="text/csv",key="ai_dl_"+str(_ex_msgs.index(_m)))
            if _m.get("excel_bytes") is not None:
                st.download_button("Download supplier_scorecard.xlsx",
                    data=_m["excel_bytes"],file_name="supplier_scorecard.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="ai_xl_"+str(_ex_msgs.index(_m)))
    if not _ex_msgs:
        st.markdown('<div class="chat-ai">Ask: which file for buyer report · CRITICAL alerts by region · supplier scorecard as Excel.</div>',unsafe_allow_html=True)
    _ex_pend=st.session_state.pop("_pend_ex",None)
    _ex_in=st.chat_input("Ask about exports...",key="ci_export")
    _ex_in=_ex_in or _ex_pend
    if _ex_in:
        _ex_msgs.append({"role":"user","content":_ex_in,"csv_bytes":None,"fname":None,"excel_bytes":None})
        with st.spinner("Preparing export..."):
            if _exai_ok:
                _etxt2,_ecsv,_efname,_eexcel=_exai.answer(_ex_in,na,D["sup_sc"],ns)
            else:
                _etxt2="export_ai module not loaded."
                _ecsv=None; _efname=None; _eexcel=None
        _ex_msgs.append({"role":"assistant","content":_etxt2,
            "csv_bytes":_ecsv,"fname":_efname,"excel_bytes":_eexcel})
        st.rerun()
